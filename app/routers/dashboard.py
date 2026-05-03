"""Dashboard API endpoints."""

import logging
import shutil
from pathlib import Path

import httpx
from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse, PlainTextResponse

from app.config import get, reload_config
from app.database import get_trades, get_stats, export_csv, get_today_trades, get_wallet_transactions, get_kamino_net_deposited, get_open_positions, get_all_positions, get_position_analytics, get_indicator_performance, get_backtests, insert_backtest, delete_backtest, get_kalshi_trades, get_kalshi_positions, get_kalshi_stats, insert_kamino_snapshot, compute_kamino_earnings, get_latest_kamino_snapshot, insert_portfolio_snapshot, get_portfolio_snapshots, get_latest_portfolio_snapshot
from app.models import DashboardStats, SettingsUpdate
from app.services.jupiter_client import JupiterClient
from app.services.wallet_service import WalletService
from app.services.price_feed import get_price_feed
from app import state
from app.services.kamino_client import KaminoClient
from app.services.ngrok_monitor import get_ngrok_monitor
from app.services.paper_trading import get_paper_trader
from app.services.scout_service import get_latest_report, get_all_reports, ScoutService

logger = logging.getLogger("bot.dashboard")

router = APIRouter(prefix="/api", tags=["dashboard"])


# Module-level wallet singleton — keeps the 15s balance cache alive across requests
_wallet: WalletService = None

def _get_wallet() -> WalletService:
    global _wallet
    if _wallet is None:
        _wallet = WalletService()
    return _wallet


# Module-level EVM wallet singleton (lazy — only loads if config has evm_wallet block)
_evm_wallet = None
_evm_wallet_unavailable = False  # cache "no config" state to skip retries

def _get_evm_wallet():
    """Lazy-load the EVM wallet. Returns None if config missing — dashboard
    should gracefully render an EVM section as 'not configured' in that case."""
    global _evm_wallet, _evm_wallet_unavailable
    if _evm_wallet_unavailable:
        return None
    if _evm_wallet is None:
        try:
            from app.services.evm_wallet_service import EVMWalletService
            _evm_wallet = EVMWalletService()
        except Exception as e:
            logger.info(f"EVM wallet unavailable (probably no config): {e}")
            _evm_wallet_unavailable = True
            return None
    return _evm_wallet


@router.get("/evm/balance")
async def get_evm_balance():
    """EVM wallet balances on Arbitrum (or other configured chain).

    Returns ETH (gas), USDC, and any non-zero held tokens with USD valuations
    where prices are available (currently ETH only — token USD valuations are
    a follow-up).
    """
    wallet = _get_evm_wallet()
    if wallet is None:
        return {
            "configured": False,
            "address": None,
            "subtotal_usd": 0.0,
            "errors": {"config": "evm_wallet block not in config.yaml"},
        }

    out = {
        "configured": True,
        "address": wallet.address,
        "chain_id": wallet.chain_id,
        "chain": "arbitrum" if wallet.chain_id == 42161 else f"chain_{wallet.chain_id}",
        "eth_native": 0.0,
        "eth_usd": 0.0,
        "usdc": 0.0,
        "token_holdings": {},
        "tokens_usd": 0.0,
        "subtotal_usd": 0.0,
        "errors": {},
    }

    # ETH price: try price_feed first (may not have ETH), fall back to a known reference
    eth_price = 0.0
    try:
        feed = get_price_feed()
        if feed.is_running:
            all_prices = feed.get_all_prices()
            eth_price = (all_prices.get("ETH") or {}).get("price", 0.0)
    except Exception:
        pass
    # Fallback: use a recent price hint via Jupiter (Solana wrapped ETH proxy)
    if eth_price <= 0:
        try:
            jupiter = JupiterClient()
            try:
                eth_price = await jupiter.get_token_price("ETH") or 0.0
            finally:
                await jupiter.close()
        except Exception as e:
            out["errors"]["eth_price"] = str(e)[:120]

    # Fetch native ETH + USDC + tracked token balances in parallel
    import asyncio as _a
    try:
        eth_bal, usdc_bal, holdings = await _a.gather(
            wallet.get_eth_balance(),
            wallet.get_usdc_balance(),
            wallet.get_tracked_token_balances(),
        )
        out["eth_native"] = eth_bal
        out["eth_usd"] = eth_bal * eth_price
        out["usdc"] = usdc_bal
        out["token_holdings"] = {
            sym: {"amount": amt, "price_usd": 0.0, "usd_value": 0.0}
            for sym, amt in holdings.items()
        }
        # USD valuation for held tokens — best-effort via price_feed (Solana feed
        # may have wrapped versions of some EVM tokens, e.g. RENDER, JTO).
        # Tokens without prices show as 0 USD value but their amounts still display.
        try:
            feed = get_price_feed()
            if feed.is_running:
                all_prices = feed.get_all_prices()
                for sym, info in out["token_holdings"].items():
                    p = (all_prices.get(sym) or {}).get("price", 0.0)
                    if p > 0:
                        info["price_usd"] = p
                        info["usd_value"] = info["amount"] * p
                        out["tokens_usd"] += info["usd_value"]
        except Exception:
            pass
        out["subtotal_usd"] = out["eth_usd"] + out["usdc"] + out["tokens_usd"]
    except Exception as e:
        logger.warning(f"/evm/balance fetch error: {e}")
        out["errors"]["balance"] = str(e)[:200]

    return out


@router.get("/stats")
async def get_dashboard_stats():
    """Get live dashboard statistics."""
    try:
        wallet = _get_wallet()

        # Use PriceFeed (real-time WebSocket) for prices — instant, no network call
        feed = get_price_feed()
        token_prices = {}
        sol_price = 0.0
        if feed.is_running:
            all_prices = feed.get_all_prices()
            if "SOL" in all_prices:
                sol_price = all_prices["SOL"]["price"]
            # Convert to format wallet expects: {symbol: {"price": float}}
            token_prices = all_prices

        # Fallback to HTTP if PriceFeed has no SOL price yet
        if sol_price <= 0:
            jupiter = JupiterClient()
            try:
                sol_price = await jupiter.get_sol_price()
                if not token_prices:
                    token_prices = await jupiter.get_multi_token_prices()
            finally:
                await jupiter.close()

        balances = await wallet.get_total_usd_balance(sol_price, token_prices)

        db_stats = get_stats()
        total = db_stats["total_trades"]
        wins = db_stats["winning_trades"]
        total_usd = balances["total_usd"]

        return DashboardStats(
            wallet_balance_sol=balances["sol"],
            wallet_balance_usdc=balances["usdc"],
            wallet_balance_usd=total_usd,
            sol_price_usd=sol_price,
            token_holdings=balances.get("token_holdings", {}),
            tokens_usd=balances.get("tokens_usd", 0.0),
            total_trades=total,
            winning_trades=wins,
            losing_trades=db_stats["losing_trades"],
            win_rate=(wins / total * 100) if total > 0 else 0,
            total_pnl_usd=db_stats["total_pnl_usd"],
            total_pnl_percent=(db_stats["total_pnl_usd"] / total_usd * 100) if total_usd > 0 else 0,
            today_pnl_usd=db_stats["today_pnl_usd"],
            avg_trade_size_sol=db_stats["avg_trade_size_sol"],
            last_signal_time=db_stats["last_trade_time"],
            bot_status="running",
        )
    except Exception as e:
        logger.error(f"Stats error: {e}")
        db_stats = get_stats()
        return DashboardStats(
            total_trades=db_stats["total_trades"],
            total_pnl_usd=db_stats["total_pnl_usd"],
            today_pnl_usd=db_stats["today_pnl_usd"],
            bot_status="error",
        )


@router.get("/portfolio")
async def get_unified_portfolio():
    """Unified portfolio view: crypto wallet + Kamino + Kalshi totals with P&L.

    Also records a portfolio snapshot (rate-limited to 1/10min) for the equity curve.
    """
    import asyncio as _a
    from datetime import datetime as _dt

    out: dict = {
        "timestamp": _dt.utcnow().isoformat(),
        "crypto": {"sol": 0.0, "usdc": 0.0, "tokens_usd": 0.0, "subtotal_usd": 0.0},
        "evm":    {"address": None, "configured": False, "eth_native": 0.0, "usdc": 0.0,
                   "tokens_usd": 0.0, "subtotal_usd": 0.0},
        "kamino": {"deposited_usd": 0.0, "earnings_usd": 0.0, "supply_apy": 0.0, "daily_yield_est": 0.0},
        "kalshi": {"balance_usd": 0.0, "invested_usd": 0.0, "pnl_usd": 0.0, "open_positions": 0},
        "totals": {"total_usd": 0.0, "today_pnl_usd": 0.0, "all_time_pnl_usd": 0.0, "delta_24h_usd": 0.0, "delta_24h_pct": 0.0},
        "errors": {},
    }

    # --- Crypto wallet via existing stats pathway ---
    try:
        wallet = _get_wallet()
        feed = get_price_feed()
        token_prices, sol_price = {}, 0.0
        if feed.is_running:
            all_prices = feed.get_all_prices()
            sol_price = all_prices.get("SOL", {}).get("price", 0.0)
            token_prices = all_prices
        if sol_price <= 0:
            jupiter = JupiterClient()
            try:
                sol_price = await jupiter.get_sol_price()
                if not token_prices:
                    token_prices = await jupiter.get_multi_token_prices()
            finally:
                await jupiter.close()

        balances = await wallet.get_total_usd_balance(sol_price, token_prices)
        out["crypto"] = {
            "sol": balances["sol"],
            "sol_usd": balances["sol_usd_value"],
            "usdc": balances["usdc"],
            "tokens_usd": balances["tokens_usd"],
            "token_holdings": balances.get("token_holdings", {}),
            "subtotal_usd": balances["total_usd"],
            "sol_price_usd": sol_price,
        }
    except Exception as e:
        logger.warning(f"/portfolio crypto error: {e}")
        out["errors"]["crypto"] = str(e)[:160]

    # --- Kamino ---
    kamino_usd = 0.0
    try:
        kamino = KaminoClient()
        wallet_for_kamino = _get_wallet()
        try:
            metrics = await kamino.get_reserve_metrics()
            position = await kamino.get_user_position(wallet_for_kamino.public_key)
        finally:
            await kamino.close()
        deposited = position.get("deposited_usdc", 0)
        kamino_usd = deposited
        earnings = compute_kamino_earnings(current_balance=deposited)
        supply_apy = metrics.get("supply_apy", 0)
        out["kamino"] = {
            "deposited_usd": deposited,
            "earnings_usd": earnings["earnings_total"],
            "supply_apy": supply_apy,
            "daily_yield_est": deposited * supply_apy / 100 / 365,
            "monthly_yield_est": deposited * supply_apy / 100 / 12,
            "stale": position.get("stale", False),
        }
    except Exception as e:
        logger.warning(f"/portfolio kamino error: {e}")
        out["errors"]["kamino"] = str(e)[:160]

    # --- Kalshi (balance + live open-position market value via API; realized P&L via DB) ---
    kalshi_cash_usd = 0.0
    kalshi_open_value_usd = 0.0
    kalshi_unrealized_pnl_usd = 0.0
    kalshi_live_open = 0
    try:
        from app.services.kalshi_client import get_kalshi_client, AsyncKalshiClient
        kcli = get_kalshi_client()
        akc = AsyncKalshiClient(kcli)

        # Cash balance
        try:
            bal = await akc.get_balance()
            kalshi_cash_usd = int((bal or {}).get("balance", 0) or 0) / 100
        except Exception as e:
            out["errors"]["kalshi_balance"] = str(e)[:160]

        # Live open positions (market value + unrealized P&L)
        try:
            summ = await akc.get_portfolio_summary()
            kalshi_open_value_usd = (summ.get("total_market_value_cents", 0) or 0) / 100
            invested_open_usd = (summ.get("total_invested_cents", 0) or 0) / 100
            kalshi_unrealized_pnl_usd = (summ.get("unrealized_pnl_cents", 0) or 0) / 100
            # Count positions with non-zero count
            kalshi_live_open = sum(
                1 for p in (summ.get("positions", []) or [])
                if (p.get("count") or 0) > 0
            )
        except Exception as e:
            out["errors"]["kalshi_positions"] = str(e)[:160]
            invested_open_usd = 0.0

        # Realized P&L from local DB (closed Kalshi trades — captures only positions
        # opened after we started recording; misses pre-bot settlement history).
        kstats = get_kalshi_stats()
        realized_pnl_usd = (kstats.get("total_pnl_cents", 0) or 0) / 100

        # All-time Kalshi P&L = current_value - starting_capital. Kalshi's settlement
        # endpoint isn't pulled into the local DB, so we anchor against config-recorded
        # cumulative deposits (kalshi.starting_capital_usd, set 2026-04-29).
        kalshi_current_total = kalshi_cash_usd + kalshi_open_value_usd
        starting_capital_usd = float(get("kalshi", "starting_capital_usd", 0.0) or 0.0)
        kalshi_alltime_pnl_usd = (
            kalshi_current_total - starting_capital_usd
            if starting_capital_usd > 0
            else realized_pnl_usd + kalshi_unrealized_pnl_usd
        )

        out["kalshi"] = {
            "balance_usd": kalshi_cash_usd,
            "open_value_usd": kalshi_open_value_usd,
            "invested_usd": invested_open_usd,
            "unrealized_pnl_usd": kalshi_unrealized_pnl_usd,
            "realized_pnl_usd": realized_pnl_usd,
            # `pnl_usd` now reflects all-time P&L vs deposits, not just recent DB rows.
            "pnl_usd": kalshi_alltime_pnl_usd,
            "starting_capital_usd": starting_capital_usd,
            "total_value_usd": kalshi_current_total,
            "open_positions": kalshi_live_open or (kstats.get("open_positions", 0) or 0),
            "total_positions": kstats.get("total_positions", 0) or 0,
            "winning": kstats.get("winning", 0) or 0,
            "losing": kstats.get("losing", 0) or 0,
        }
    except Exception as e:
        logger.warning(f"/portfolio kalshi error: {e}")
        out["errors"]["kalshi"] = str(e)[:160]

    # --- EVM wallet (Arbitrum) ---
    evm_usd = 0.0
    try:
        evm_data = await get_evm_balance()
        if evm_data.get("configured"):
            out["evm"] = {
                "address":     evm_data.get("address"),
                "configured":  True,
                "chain":       evm_data.get("chain"),
                "eth_native":  evm_data.get("eth_native", 0.0),
                "eth_usd":     evm_data.get("eth_usd", 0.0),
                "usdc":        evm_data.get("usdc", 0.0),
                "tokens_usd":  evm_data.get("tokens_usd", 0.0),
                "token_holdings": evm_data.get("token_holdings", {}),
                "subtotal_usd": evm_data.get("subtotal_usd", 0.0),
            }
            evm_usd = evm_data.get("subtotal_usd", 0.0)
    except Exception as e:
        logger.warning(f"/portfolio evm error: {e}")
        out["errors"]["evm"] = str(e)[:160]

    # --- Totals ---
    crypto_usd = out["crypto"]["subtotal_usd"]
    # Kalshi contribution = cash + live market value of open contracts
    kalshi_usd = kalshi_cash_usd + kalshi_open_value_usd
    total_usd = crypto_usd + evm_usd + kamino_usd + kalshi_usd
    db_stats = get_stats()
    all_time_pnl = db_stats.get("total_pnl_usd", 0.0) + out["kalshi"].get("pnl_usd", 0.0)
    today_pnl = db_stats.get("today_pnl_usd", 0.0)

    # 24h delta vs snapshot from ~24h ago
    delta_24h_usd, delta_24h_pct = 0.0, 0.0
    try:
        import sqlite3
        from app.database import DB_PATH
        c = sqlite3.connect(DB_PATH)
        c.row_factory = sqlite3.Row
        row = c.execute(
            """SELECT total_usd FROM portfolio_snapshots
               WHERE timestamp <= datetime('now','-23 hours')
               ORDER BY timestamp DESC LIMIT 1"""
        ).fetchone()
        c.close()
        if row and row["total_usd"]:
            delta_24h_usd = total_usd - row["total_usd"]
            delta_24h_pct = (delta_24h_usd / row["total_usd"]) * 100 if row["total_usd"] > 0 else 0.0
    except Exception as e:
        logger.debug(f"24h delta lookup failed: {e}")

    out["totals"] = {
        "total_usd": total_usd,
        "crypto_usd": crypto_usd,
        "evm_usd": evm_usd,
        "kamino_usd": kamino_usd,
        "kalshi_usd": kalshi_usd,
        "today_pnl_usd": today_pnl,
        "all_time_pnl_usd": all_time_pnl,
        "delta_24h_usd": delta_24h_usd,
        "delta_24h_pct": delta_24h_pct,
    }

    # Snapshot (rate-limited): max one every 5 min
    try:
        latest = get_latest_portfolio_snapshot()
        write_snap = True
        if latest:
            last_ts = _dt.fromisoformat(latest["timestamp"]).timestamp()
            import time as _t
            if _t.time() - last_ts < 300:
                write_snap = False
        if write_snap and total_usd > 0:
            insert_portfolio_snapshot({
                "sol_usd": out["crypto"].get("sol_usd", 0.0),
                "usdc_usd": out["crypto"].get("usdc", 0.0),
                "tokens_usd": out["crypto"].get("tokens_usd", 0.0),
                "kamino_usd": kamino_usd,
                "kalshi_usd": kalshi_usd,
                "total_usd": total_usd,
            })
    except Exception as e:
        logger.debug(f"portfolio snapshot write failed: {e}")

    return out


@router.get("/wallet/tokens")
async def get_wallet_tokens():
    """Per-token list with target/current/drift %, holding, avg entry, unrealized P&L.

    Used by the dashboard wallet token list (replaces the legacy tile grid).
    """
    out: dict = {"tokens": [], "total_portfolio_usd": 0.0, "errors": {}}

    # --- Resolve prices + holdings (same pattern as /stats) ---
    try:
        wallet = _get_wallet()
        feed = get_price_feed()
        token_prices: dict = {}
        sol_price = 0.0
        if feed.is_running:
            all_prices = feed.get_all_prices()
            sol_price = all_prices.get("SOL", {}).get("price", 0.0)
            token_prices = all_prices
        if sol_price <= 0:
            jupiter = JupiterClient()
            try:
                sol_price = await jupiter.get_sol_price()
                if not token_prices:
                    token_prices = await jupiter.get_multi_token_prices()
            finally:
                await jupiter.close()
        balances = await wallet.get_total_usd_balance(sol_price, token_prices)
    except Exception as e:
        logger.error(f"/wallet/tokens balances error: {e}")
        out["errors"]["balances"] = str(e)[:160]
        return out

    total_usd = float(balances.get("total_usd") or 0.0)
    out["total_portfolio_usd"] = total_usd
    token_holdings = balances.get("token_holdings") or {}

    # --- Targets from rebalancer config ---
    targets = get("rebalancer", "targets", {}) or {}

    # --- Cost basis from open positions (per symbol, weighted avg entry) ---
    cost_basis: dict = {}
    try:
        for p in get_open_positions():
            sym = p.get("symbol")
            amt_usdc = float(p.get("amount_usdc") or 0)
            entry = float(p.get("entry_price") or 0)
            if sym and amt_usdc > 0 and entry > 0:
                tokens = amt_usdc / entry
                cb = cost_basis.setdefault(sym, {"total_usd": 0.0, "total_tokens": 0.0})
                cb["total_usd"] += amt_usdc
                cb["total_tokens"] += tokens
    except Exception as e:
        logger.debug(f"/wallet/tokens cost-basis error: {e}")

    # --- Universe: held + targeted + priced (skip USDC; shown separately on dashboard) ---
    sol_holding = {
        "amount": float(balances.get("sol") or 0.0),
        "price": sol_price,
        "usd_value": float(balances.get("sol_usd_value") or 0.0),
        "chain": "solana",
    }
    # Mark Solana SPL holdings with chain info
    sol_token_holdings = {
        sym: {**info, "chain": "solana"} for sym, info in token_holdings.items()
    }
    all_holdings = {"SOL": sol_holding, **sol_token_holdings}

    # --- EVM wallet holdings (Arbitrum) — merge into the same universe ---
    evm_chain_map = {}  # symbol → "arbitrum" for badge rendering
    try:
        evm_wallet = _get_evm_wallet()
        if evm_wallet is not None:
            evm_eth = await evm_wallet.get_eth_balance()
            evm_holdings = await evm_wallet.get_tracked_token_balances()
            # ETH balance shown as a holding (gas + tradable asset)
            if evm_eth > 0:
                eth_price = (token_prices.get("ETH") or {}).get("price", 0.0)
                all_holdings["ETH"] = {
                    "amount": evm_eth,
                    "price": eth_price,
                    "usd_value": evm_eth * eth_price,
                    "chain": "arbitrum",
                }
                evm_chain_map["ETH"] = "arbitrum"
            for sym, amt in evm_holdings.items():
                price = (token_prices.get(sym) or {}).get("price", 0.0)
                # If symbol already exists from Solana, prefer the larger holding
                # but keep the EVM one with a different key suffix to avoid collisions.
                # For now, EVM-only symbols (INJ, ARB, LDO, etc.) won't collide.
                if sym in all_holdings and all_holdings[sym].get("chain") == "solana":
                    # Edge case: symbol exists on both chains. Prefix EVM one for clarity.
                    key = f"{sym}.ARB"
                else:
                    key = sym
                all_holdings[key] = {
                    "amount": amt,
                    "price": price,
                    "usd_value": amt * price if price > 0 else 0.0,
                    "chain": "arbitrum",
                }
                evm_chain_map[key] = "arbitrum"
    except Exception as e:
        logger.warning(f"/wallet/tokens EVM merge error: {e}")
        out["errors"]["evm"] = str(e)[:160]

    # Always include the bot's tracked EVM tokens (even if not held) so the user
    # can see their full trading universe at a glance — INJ/LINK/AAVE etc. show
    # as zero-balance rows until the alert fires and a position opens.
    try:
        from app.services.trade_engine import TradeEngine
        for sym in TradeEngine.EVM_TOKENS:
            if sym not in all_holdings:
                all_holdings[sym] = {"amount": 0.0, "price": 0.0,
                                     "usd_value": 0.0, "chain": "arbitrum"}
    except Exception:
        pass

    universe = set(all_holdings.keys()) | set(targets.keys()) | set(token_prices.keys())
    universe.discard("USDC")

    rows = []
    for sym in universe:
        holding = all_holdings.get(sym, {"amount": 0.0, "price": 0.0, "usd_value": 0.0})
        amount = float(holding.get("amount") or 0.0)
        pinfo = token_prices.get(sym, {}) or {}
        price = float(pinfo.get("price") or holding.get("price") or 0.0)
        change_24h = float(pinfo.get("change_24h") or 0.0)
        usd_value = amount * price if price > 0 else float(holding.get("usd_value") or 0.0)

        target_pct = targets.get(sym)
        current_pct = (usd_value / total_usd * 100) if total_usd > 0 else 0.0
        drift_pct = (current_pct - float(target_pct)) if target_pct is not None else None

        cb = cost_basis.get(sym)
        avg_entry = (cb["total_usd"] / cb["total_tokens"]) if cb and cb["total_tokens"] > 0 else None
        if avg_entry and amount > 0 and price > 0:
            tokens_in_position = min(amount, cb["total_tokens"])
            unrealized_pnl_usd = (price - avg_entry) * tokens_in_position
            unrealized_pnl_pct = ((price / avg_entry) - 1) * 100
        else:
            unrealized_pnl_usd = None
            unrealized_pnl_pct = None

        # Mark as "tracked" if it's part of the bot's deployed-alert universe,
        # even when balance is zero. Lets the dashboard show INJ/LINK/AAVE
        # as watched-but-unfilled rows so user sees the full active universe.
        is_tracked = False
        try:
            from app.services.trade_engine import TradeEngine
            if sym in TradeEngine.EVM_TOKENS:
                is_tracked = True
        except Exception:
            pass

        rows.append({
            "symbol": sym,
            "price": price,
            "change_24h": change_24h,
            "amount": amount,
            "usd_value": usd_value,
            "target_pct": float(target_pct) if target_pct is not None else None,
            "current_pct": current_pct,
            "drift_pct": drift_pct,
            "avg_entry": avg_entry,
            "unrealized_pnl_usd": unrealized_pnl_usd,
            "unrealized_pnl_pct": unrealized_pnl_pct,
            "has_position": amount > 0 and usd_value > 0.01,
            "is_tracked": is_tracked,
            "chain": all_holdings.get(sym, {}).get("chain", "solana"),
        })

    # Sort: targeted first by target desc, then untargeted by current % desc
    rows.sort(key=lambda r: (
        0 if r["target_pct"] is not None else 1,
        -(r["target_pct"] or 0),
        -r["current_pct"],
    ))
    out["tokens"] = rows
    return out


@router.get("/portfolio/equity")
async def get_portfolio_equity(days: int = 30):
    """Portfolio equity curve for the last N days (from portfolio_snapshots)."""
    days = max(1, min(int(days), 365))
    rows = get_portfolio_snapshots(days=days)
    return {
        "days": days,
        "points": [
            {
                "t": r["timestamp"],
                "total": r["total_usd"],
                "crypto": r["sol_usd"] + r["usdc_usd"] + r["tokens_usd"],
                "kamino": r["kamino_usd"],
                "kalshi": r["kalshi_usd"],
            }
            for r in rows
        ],
    }


@router.get("/strategy/pnl")
async def get_strategy_pnl(days: int = 30):
    """Per-strategy P&L for the last N days (crypto trades + Kalshi).

    Crypto side: sum of `pnl_usd` on closed trades grouped by strategy.
    Kalshi side: rolled up under a single 'kalshi' bar.
    """
    days = max(1, min(int(days), 365))
    import sqlite3
    from app.database import DB_PATH
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    rows = c.execute(
        """SELECT COALESCE(NULLIF(strategy, ''), 'manual') AS strategy,
                  COALESCE(SUM(pnl_usd), 0) AS pnl,
                  COUNT(*) AS trades
           FROM trades
           WHERE pnl_usd IS NOT NULL AND timestamp >= datetime('now', ?)
           GROUP BY COALESCE(NULLIF(strategy, ''), 'manual')
           ORDER BY pnl DESC""",
        (f"-{days} days",),
    ).fetchall()
    kalshi_pnl_cents = c.execute(
        """SELECT COALESCE(SUM(pnl_cents), 0) AS p, COUNT(*) AS n
           FROM kalshi_positions
           WHERE status != 'open' AND COALESCE(closed_at, opened_at) >= datetime('now', ?)""",
        (f"-{days} days",),
    ).fetchone()
    c.close()

    bars = [{"strategy": r["strategy"], "pnl_usd": float(r["pnl"] or 0), "trades": int(r["trades"] or 0)} for r in rows]
    kpnl = (kalshi_pnl_cents["p"] or 0) / 100
    knt = kalshi_pnl_cents["n"] or 0
    if knt > 0:
        bars.append({"strategy": "kalshi", "pnl_usd": kpnl, "trades": knt})
    bars.sort(key=lambda x: x["pnl_usd"], reverse=True)
    return {"days": days, "bars": bars}


@router.get("/trades")
async def get_trade_history(limit: int = 100, offset: int = 0):
    """Get trade history."""
    trades = get_trades(limit=min(limit, 500), offset=offset)
    return {"trades": trades, "total": len(trades)}


@router.get("/trades/today")
async def get_todays_trades():
    """Get today's trades."""
    return {"trades": get_today_trades()}


@router.get("/price")
async def get_sol_price():
    """Get current SOL price."""
    jupiter = JupiterClient()
    try:
        price = await jupiter.get_sol_price()
        market = await jupiter.get_market_data()
        return {"price": price, "market_data": market}
    finally:
        await jupiter.close()


@router.get("/prices")
async def get_token_prices():
    """Get live prices for all tracked tokens.

    Uses the real-time PriceFeed (WebSocket) when available,
    falls back to Jupiter HTTP polling otherwise.
    """
    from app.services.price_feed import get_price_feed

    feed = get_price_feed()
    if feed.is_running:
        prices = feed.get_all_prices()
        if prices:
            return {"prices": prices}

    # Fallback: HTTP polling via Jupiter
    jupiter = JupiterClient()
    try:
        prices = await jupiter.get_multi_token_prices()
        return {"prices": prices}
    except Exception as e:
        logger.error(f"Multi-token price error: {e}")
        return {"prices": {}}
    finally:
        await jupiter.close()


@router.get("/settings")
async def get_settings():
    """Get current risk/trading settings."""
    risk = get("risk")
    geo = get("geo_risk")
    jupiter = get("jupiter")
    return {
        "risk": risk,
        "geo_risk": geo,
        "jupiter": {
            "slippage_bps": jupiter.get("slippage_bps", 100),
            "priority_fee_lamports": jupiter.get("priority_fee_lamports", 50000),
        },
    }


@router.post("/settings")
async def update_settings(updates: SettingsUpdate):
    """Update risk/trading settings (runtime only, does not persist to YAML)."""
    # For production, you'd write back to config.yaml
    # This updates the in-memory config
    import yaml

    config_path = Path("config.yaml")
    if not config_path.exists():
        raise HTTPException(status_code=500, detail="config.yaml not found")

    with open(config_path) as f:
        cfg = yaml.safe_load(f)

    update_dict = updates.model_dump(exclude_none=True)
    for key, val in update_dict.items():
        if key == "geo_risk_weight":
            cfg.setdefault("geo_risk", {})["weight"] = val
        elif key in ("slippage_bps", "priority_fee_lamports"):
            cfg.setdefault("jupiter", {})[key] = val
        elif key in cfg.get("risk", {}):
            cfg["risk"][key] = val

    with open(config_path, "w") as f:
        yaml.dump(cfg, f, default_flow_style=False)

    reload_config()
    return {"status": "ok", "updated": update_dict}


@router.get("/export/csv")
async def export_trades_csv():
    """Export all trades as CSV for tax compliance."""
    path = export_csv()
    if not Path(path).exists():
        raise HTTPException(status_code=404, detail="No trades to export")
    return FileResponse(path, media_type="text/csv", filename=Path(path).name)


@router.get("/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "service": "tradingview-bot"}


@router.get("/monitor/alerts")
async def monitor_alerts():
    """Return latest monitor alert and log tail for the dashboard."""
    import json
    bot_dir = Path(__file__).resolve().parent.parent.parent
    alert_file = bot_dir / "logs" / "monitor_alert.json"
    monitor_log = bot_dir / "logs" / "monitor.log"
    bot_log = bot_dir / "logs" / "bot.log"

    result = {"has_alerts": False, "issues": [], "timestamp": None, "recent_errors": [], "monitor_log": []}

    # Read alert file if present
    if alert_file.exists():
        try:
            data = json.loads(alert_file.read_text())
            result["has_alerts"] = True
            result["issues"] = data.get("issues", [])
            result["timestamp"] = data.get("timestamp")
        except Exception:
            pass

    # Tail monitor log (last 10 lines)
    if monitor_log.exists():
        try:
            lines = monitor_log.read_text().strip().splitlines()
            result["monitor_log"] = lines[-10:]
        except Exception:
            pass

    # Recent errors from bot.log (last 20 ERROR lines)
    if bot_log.exists():
        try:
            errors = []
            for line in reversed(bot_log.read_text().splitlines()):
                if "[ERROR]" in line:
                    errors.append(line.strip()[:200])
                    if len(errors) >= 20:
                        break
            result["recent_errors"] = errors
        except Exception:
            pass

    return result


@router.get("/bot/status")
async def bot_status():
    """Bot running state + uptime."""
    return {
        "active": state.is_active(),
        "uptime": state.get_uptime(),
    }


@router.post("/bot/start")
async def bot_start():
    return state.start_bot()


@router.post("/bot/stop")
async def bot_stop():
    return state.stop_bot()


@router.get("/system/status")
async def system_status():
    """Check all component connections and return status map."""
    results = {}

    def err(e): return str(e)[:60]

    # 1. Database
    try:
        from app.database import get_db
        conn = get_db()
        conn.execute("SELECT 1").fetchone()
        conn.close()
        results["database"] = {"ok": True, "label": "SQLite"}
    except Exception as e:
        results["database"] = {"ok": False, "label": "SQLite", "error": err(e)}

    # 2. Wallet / RPC
    try:
        wallet = WalletService()
        bal = await wallet.get_balance_sol()
        await wallet.close()
        results["wallet"] = {"ok": True, "label": "Solana RPC", "detail": f"{bal:.4f} SOL"}
    except Exception as e:
        results["wallet"] = {"ok": False, "label": "Solana RPC", "error": err(e)}

    # 3. Price feed — Binance public API (no key, no rate limit)
    try:
        async with httpx.AsyncClient(timeout=8) as client:
            resp = await client.get(
                "https://api.binance.us/api/v3/ticker/price",
                params={"symbol": "SOLUSDT"},
            )
            resp.raise_for_status()
            price = float(resp.json()["price"])
            results["price_feed"] = {"ok": True, "label": "Binance US", "detail": f"SOL ${price:.2f}"}
    except Exception as e:
        results["price_feed"] = {"ok": False, "label": "Binance Feed", "error": str(e)[:60]}

    # 4. Jupiter Swap API (quote endpoint ping)
    try:
        async with httpx.AsyncClient(timeout=8) as client:
            resp = await client.get(
                    "https://lite-api.jup.ag/swap/v1/quote",
                    params={
                        "inputMint": "So11111111111111111111111111111111111111112",
                        "outputMint": "EPjFWdd5AufqSSqeM2qN1xzybapC8G4wEGGkZwyTDt1v",
                        "amount": "1000000",
                        "slippageBps": "50",
                    },
                )
            results["jupiter"] = {"ok": resp.status_code < 500, "label": "Jupiter DEX", "detail": f"HTTP {resp.status_code}"}
    except Exception as e:
        results["jupiter"] = {"ok": False, "label": "Jupiter DEX", "error": err(e)}

    # 5. Claude CLI
    cli_path = get("claude", "cli_path", "claude")
    cli_mode = get("claude", "mode", "cli")
    if cli_mode == "cli":
        found = shutil.which(cli_path)
        results["claude"] = {
            "ok": found is not None,
            "label": "Claude CLI",
            "detail": found or "not found in PATH",
        }
    else:
        api_key = get("claude", "api_key", "")
        results["claude"] = {
            "ok": bool(api_key),
            "label": "Claude API",
            "detail": "key configured" if api_key else "no api_key set",
        }

    # 6. Telegram
    tg_cfg = get("telegram")
    tg_token = tg_cfg.get("bot_token", "")
    tg_enabled = tg_cfg.get("enabled", False)
    if tg_enabled and tg_token:
        try:
            async with httpx.AsyncClient(timeout=8) as client:
                resp = await client.get(f"https://api.telegram.org/bot{tg_token}/getMe")
                data = resp.json()
                ok = data.get("ok", False)
                name = data.get("result", {}).get("username", "?") if ok else "auth failed"
                results["telegram"] = {"ok": ok, "label": "Telegram Bot", "detail": f"@{name}"}
        except Exception as e:
            results["telegram"] = {"ok": False, "label": "Telegram Bot", "error": err(e)}
    else:
        results["telegram"] = {"ok": False, "label": "Telegram Bot", "detail": "disabled or no token"}

    # 7. News providers (with fallback)
    news_cfg = get("news")
    newsapi_key = news_cfg.get("newsapi_key", "")
    tavily_key = news_cfg.get("tavily_api_key", "")
    preferred = news_cfg.get("provider", "newsapi")
    news_ok = False

    async with httpx.AsyncClient(timeout=8) as client:
        # Try preferred provider first
        providers = []
        if preferred == "newsapi":
            if newsapi_key:
                providers.append(("NewsAPI", "https://newsapi.org/v2/top-headlines",
                                  {"country": "us", "pageSize": 1, "apiKey": newsapi_key}))
            if tavily_key:
                providers.append(("Tavily", "https://api.tavily.com/search", None))
        else:
            if tavily_key:
                providers.append(("Tavily", "https://api.tavily.com/search", None))
            if newsapi_key:
                providers.append(("NewsAPI", "https://newsapi.org/v2/top-headlines",
                                  {"country": "us", "pageSize": 1, "apiKey": newsapi_key}))

        for name, url, params in providers:
            try:
                if name == "Tavily":
                    resp = await client.post(url, json={
                        "api_key": tavily_key, "query": "crypto", "max_results": 1, "search_depth": "basic"
                    })
                else:
                    resp = await client.get(url, params=params)
                if resp.status_code == 200:
                    results["news"] = {"ok": True, "label": name, "detail": "active"}
                    news_ok = True
                    break
                else:
                    results["news"] = {"ok": False, "label": name, "detail": f"HTTP {resp.status_code}"}
            except Exception as e:
                results["news"] = {"ok": False, "label": name, "error": err(e)}

        if not providers:
            results["news"] = {"ok": False, "label": "News", "detail": "no keys configured"}

    # 8. Kamino Lend
    kamino_cfg = get("kamino")
    if kamino_cfg.get("enabled"):
        try:
            kamino = KaminoClient()
            metrics = await kamino.get_reserve_metrics()
            await kamino.close()
            if metrics.get("available"):
                results["kamino"] = {
                    "ok": True,
                    "label": "Kamino Lend",
                    "detail": f"USDC {metrics['supply_apy']:.2f}% APY",
                }
            else:
                results["kamino"] = {"ok": False, "label": "Kamino Lend", "detail": metrics.get("error", "unavailable")}
        except Exception as e:
            results["kamino"] = {"ok": False, "label": "Kamino Lend", "error": err(e)}
    else:
        results["kamino"] = {"ok": False, "label": "Kamino Lend", "detail": "disabled"}

    # 9. ngrok tunnel
    ngrok = get_ngrok_monitor()
    ngrok_status = ngrok.get_status()
    results["ngrok"] = {
        "ok": ngrok_status["online"],
        "label": "ngrok Tunnel",
        "detail": ngrok_status["current_url"] or "not detected",
    }

    return {
        "components": results,
        "bot_active": state.is_active(),
        "uptime": state.get_uptime(),
        "ngrok": ngrok_status,
    }


@router.get("/ngrok")
async def get_ngrok_status():
    """Get current ngrok tunnel status with URL history."""
    ngrok = get_ngrok_monitor()
    status = await ngrok.check_once()
    return status


@router.get("/kamino")
async def get_kamino_status():
    """Get Kamino Lend yield status and position info."""
    import time as _t
    kamino = KaminoClient()
    wallet = WalletService()

    try:
        metrics = await kamino.get_reserve_metrics()
        position = await kamino.get_user_position(wallet.public_key)
        usdc_balance = await wallet.get_usdc_balance()

        deposited = position.get("deposited_usdc", 0)
        position_error = position.get("error")
        position_stale = position.get("stale", False)
        total_usdc = deposited + usdc_balance
        supply_apy = metrics.get("supply_apy", 0)

        # Snapshot-based earnings tracking: rate limit to 1/5min to avoid DB churn.
        latest_snap = get_latest_kamino_snapshot()
        now = _t.time()
        should_snap = True
        if latest_snap:
            from datetime import datetime as _dt
            last_ts = _dt.fromisoformat(latest_snap["timestamp"]).timestamp()
            if now - last_ts < 300:
                should_snap = False
        if should_snap and not position_error:
            insert_kamino_snapshot(deposited, supply_apy=supply_apy, source="dashboard")

        earnings_info = compute_kamino_earnings(current_balance=deposited)

        return {
            "enabled": kamino.enabled,
            "auto_deposit": kamino.auto_deposit,
            "auto_withdraw": kamino.auto_withdraw,
            "wallet_address": wallet.public_key,
            "deposited_usdc": deposited,
            "wallet_usdc": usdc_balance,
            "total_usdc": total_usdc,
            "yield_pct_deposited": (deposited / total_usdc * 100) if total_usdc > 0 else 0,
            "supply_apy": supply_apy,
            "borrow_apy": metrics.get("borrow_apy", 0),
            "utilization": metrics.get("utilization", 0),
            "daily_yield_est": deposited * supply_apy / 100 / 365,
            "monthly_yield_est": deposited * supply_apy / 100 / 12,
            "earnings_usdc": earnings_info["earnings_total"],
            "earnings_raw_usdc": earnings_info["raw_earnings"],
            "earnings_data_quality": earnings_info["data_quality"],
            "earnings_tracked_since": earnings_info["first"],
            "position_stale": position_stale,
            "position_error": position_error,
        }
    finally:
        await kamino.close()
        await wallet.close()


@router.post("/kamino/deposit")
async def kamino_deposit(body: dict):
    """Manually deposit USDC into Kamino. Body: {"amount": 50.0} or {"percent": 80}"""
    kamino = KaminoClient()
    wallet = WalletService()

    try:
        usdc_balance = await wallet.get_usdc_balance()
        amount = body.get("amount")
        percent = body.get("percent")

        if percent is not None:
            amount = usdc_balance * float(percent) / 100

        if amount is None or amount <= 0:
            raise HTTPException(400, "Provide 'amount' (USD) or 'percent' (of wallet USDC)")

        amount = min(amount, usdc_balance - kamino.reserve_usdc)
        if amount < kamino.min_deposit:
            raise HTTPException(400, f"Amount ${amount:.2f} below minimum ${kamino.min_deposit}")

        result = await kamino.deposit(wallet.get_keypair(), amount)
        return result
    finally:
        await kamino.close()
        await wallet.close()


@router.get("/wallet/transactions")
async def get_wallet_tx_log(limit: int = 50):
    """Get wallet transaction log."""
    txs = get_wallet_transactions(limit=min(limit, 200))
    return {"transactions": txs, "total": len(txs)}


@router.post("/kamino/withdraw")
async def kamino_withdraw(body: dict):
    """Manually withdraw USDC from Kamino. Body: {"amount": 50.0} or {"percent": 100} or {"all": true}"""
    kamino = KaminoClient()
    wallet = WalletService()

    try:
        if body.get("all"):
            result = await kamino.withdraw_all(wallet.get_keypair())
            return result

        position = await kamino.get_user_position(wallet.public_key)
        deposited = position.get("deposited_usdc", 0)

        if deposited <= 0:
            raise HTTPException(400, "No USDC deposited in Kamino")

        amount = body.get("amount")
        percent = body.get("percent")

        if percent is not None:
            amount = deposited * float(percent) / 100

        if amount is None or amount <= 0:
            raise HTTPException(400, "Provide 'amount' (USD), 'percent' (of deposited), or 'all': true")

        amount = min(amount, deposited)
        result = await kamino.withdraw(wallet.get_keypair(), amount)
        return result
    finally:
        await kamino.close()
        await wallet.close()


@router.get("/scout")
async def get_scout_report():
    """Get latest scout report and recent history."""
    latest = get_latest_report()
    history = get_all_reports(limit=7)
    return {
        "latest": latest,
        "history_dates": [r.get("date") for r in history],
        "total_reports": len(history),
        "sources": latest.get("sources_searched", []) if latest else [],
    }


@router.post("/scout/run")
async def run_scout_now():
    """Manually trigger a scout scan."""
    scout = ScoutService()
    try:
        report = await scout.generate_report()
        return {"status": "ok", "result_count": report.get("result_count", 0), "report": report}
    finally:
        await scout.close()


@router.get("/positions")
async def get_positions_api(status: str = "all", limit: int = 50):
    """Get positions, optionally filtered by status."""
    if status == "open":
        positions = get_open_positions()
    else:
        positions = get_all_positions(limit=min(limit, 200))

    # Add live P&L for open positions
    has_open = any(p["status"] == "open" for p in positions)
    if has_open:
        # Use PriceFeed (real-time WebSocket) — no network call needed
        feed = get_price_feed()
        token_prices = {}
        sol_price = 0.0
        if feed.is_running:
            all_prices = feed.get_all_prices()
            if "SOL" in all_prices:
                sol_price = all_prices["SOL"]["price"]
            token_prices = all_prices

        # Fallback to HTTP if PriceFeed unavailable
        if sol_price <= 0:
            jupiter = JupiterClient()
            try:
                token_prices = await jupiter.get_multi_token_prices()
                sol_price = await jupiter.get_sol_price()
            finally:
                await jupiter.close()

        for p in positions:
            if p["status"] == "open":
                token_sym = p["symbol"].replace("USDT", "").replace("USD", "")
                if token_sym == "SOL":
                    current_price = sol_price
                elif token_sym in token_prices:
                    current_price = token_prices[token_sym].get("price", p["entry_price"])
                else:
                    current_price = p["entry_price"]
                p["current_price"] = current_price
                p["unrealized_pnl_usdc"] = (current_price - p["entry_price"]) * p["amount_sol"]
                p["unrealized_pnl_percent"] = ((current_price - p["entry_price"]) / p["entry_price"]) * 100 if p["entry_price"] > 0 else 0
                p["distance_to_tp_percent"] = ((p["tp_price"] - current_price) / current_price) * 100 if current_price > 0 else 0
                p["distance_to_sl_percent"] = ((current_price - p["sl_price"]) / current_price) * 100 if current_price > 0 else 0
                # Effective SL considers trailing stop
                trail_sl = p.get("trail_sl_price") or 0
                p["effective_sl"] = max(p["sl_price"], trail_sl)
                p["trail_active"] = trail_sl > p["sl_price"]

    return {"positions": positions, "total": len(positions)}


@router.post("/positions/{position_id}/close")
async def manual_close_position(position_id: int):
    """Manually close a position at market price."""
    from app.services.position_monitor import get_position_monitor

    positions = get_open_positions()
    pos = next((p for p in positions if p["id"] == position_id), None)
    if not pos:
        raise HTTPException(status_code=404, detail=f"Open position #{position_id} not found")

    monitor = get_position_monitor()
    jupiter = JupiterClient()
    try:
        token_sym = pos["symbol"].replace("USDT", "").replace("USD", "")
        if token_sym == "SOL":
            current_price = await jupiter.get_sol_price()
        else:
            current_price = await jupiter.get_token_price(token_sym)
        await monitor._close_position(pos, current_price, "manual", jupiter)
    finally:
        await jupiter.close()

    return {"status": "ok", "position_id": position_id}


@router.get("/positions/analytics")
async def position_analytics():
    """Get aggregated position analytics: win rate by strategy, equity curve, monthly P&L."""
    return get_position_analytics()


@router.get("/indicators/performance")
async def indicator_performance():
    """Per-indicator performance: trades, open positions, P&L, timeline."""
    return get_indicator_performance()


# ── Backtest Tracker ──────────────────────────────────────────────

@router.get("/backtests")
async def list_backtests(strategy: str = None, limit: int = 50):
    """Get backtest results, optionally filtered by strategy."""
    results = get_backtests(strategy=strategy, limit=limit)
    return {"backtests": results, "total": len(results)}


@router.post("/backtests")
async def add_backtest(data: dict):
    """Add a backtest result manually."""
    required = ["strategy_name", "version", "timeframe", "symbol"]
    for field in required:
        if field not in data:
            raise HTTPException(status_code=400, detail=f"Missing required field: {field}")
    bt_id = insert_backtest(data)
    return {"id": bt_id, "status": "ok"}


@router.post("/backtests/import")
async def import_backtests_from_reports():
    """Scan Exports/ directory and import all .report.txt files."""
    import re
    from pathlib import Path

    exports_dir = Path("Exports")
    if not exports_dir.exists():
        return {"imported": 0, "error": "Exports/ directory not found"}

    existing = {bt.get("source_file") for bt in get_backtests(limit=500)}
    imported = 0
    errors = []

    for report_file in sorted(exports_dir.glob("*.report.txt")):
        fname = report_file.name
        if fname in existing:
            continue

        try:
            text = report_file.read_text()
            bt = _parse_report_txt(text, fname)
            if bt:
                insert_backtest(bt)
                imported += 1
        except Exception as e:
            errors.append(f"{fname}: {e}")

    # Also scan xlsx files (skip those that have a matching .report.txt already imported)
    for xlsx_file in sorted(exports_dir.glob("*.xlsx")):
        fname = xlsx_file.name
        if fname in existing:
            continue
        # Skip if a .report.txt version was already imported
        report_name = fname.replace(".xlsx", ".report.txt")
        if report_name in existing:
            continue

        try:
            bt = _parse_xlsx(xlsx_file)
            if bt and bt.get("total_trades", 0) > 0:
                insert_backtest(bt)
                imported += 1
        except Exception as e:
            errors.append(f"{fname}: {e}")

    return {"imported": imported, "errors": errors, "total_in_db": len(get_backtests(limit=500))}


@router.delete("/backtests/{backtest_id}")
async def remove_backtest(backtest_id: int):
    """Delete a backtest record."""
    delete_backtest(backtest_id)
    return {"status": "ok"}


# ── Strategy Builder Tools ────────────────────────────────────────

@router.post("/backtests/parse-summary")
async def parse_tv_summary(data: dict):
    """Parse pasted TradingView Strategy Tester summary text into backtest fields."""
    from app.services.strategy_builder import parse_tv_summary as parse_fn

    text = data.get("text", "")
    if not text or len(text) < 20:
        raise HTTPException(status_code=400, detail="Paste the TradingView Strategy Tester summary text")

    parsed = parse_fn(text)

    # Merge with user-provided metadata
    for field in ["strategy_name", "version", "timeframe", "symbol", "notes"]:
        if data.get(field):
            parsed[field] = data[field]

    # Auto-save if requested and required fields present
    saved_id = None
    if data.get("auto_save") and all(parsed.get(f) for f in ["strategy_name", "version", "timeframe", "symbol"]):
        parsed["created_at"] = parsed.get("created_at") or __import__("datetime").datetime.utcnow().isoformat()
        parsed["status"] = "tested"
        saved_id = insert_backtest(parsed)

    return {"parsed": parsed, "saved_id": saved_id}


@router.post("/backtests/generate-indicator")
async def generate_indicator(data: dict):
    """Generate a webhook alert indicator from a backtest strategy Pine script."""
    from app.services.strategy_builder import generate_indicator as gen_fn

    strategy_name = data.get("strategy_name", "Strategy")
    version = data.get("version", "v1.0")
    timeframe = data.get("timeframe", "4H")
    pine_code = data.get("pine_code", "")
    tokens = data.get("tokens", [])

    if not pine_code or len(pine_code) < 50:
        raise HTTPException(status_code=400, detail="Provide the backtest strategy Pine script code")

    indicator_code = gen_fn(strategy_name, version, timeframe, pine_code, tokens)
    return {"indicator_code": indicator_code, "strategy_name": strategy_name, "version": version}


@router.post("/backtests/parameter-grid")
async def parameter_grid(data: dict):
    """Generate Pine script variants by sweeping parameter ranges."""
    from app.services.strategy_builder import generate_parameter_grid

    pine_code = data.get("pine_code", "")
    param_ranges = data.get("param_ranges", {})
    strategy_name = data.get("strategy_name", "Strategy")

    if not pine_code or not param_ranges:
        raise HTTPException(status_code=400, detail="Provide pine_code and param_ranges")

    # Limit grid size to prevent abuse
    import itertools
    total_combos = 1
    for spec in param_ranges.values():
        if isinstance(spec, list):
            total_combos *= len(spec)
        elif isinstance(spec, dict):
            count = int((spec["max"] - spec["min"]) / spec["step"]) + 1
            total_combos *= count
    if total_combos > 100:
        raise HTTPException(status_code=400, detail=f"Grid too large ({total_combos} variants). Max 100.")

    variants = generate_parameter_grid(pine_code, param_ranges, strategy_name)
    return {"variants": [{"params": v["params"], "label": v["label"], "pine_code": v["pine_code"]} for v in variants], "total": len(variants)}


@router.post("/backtests/token-sweep")
async def token_sweep(data: dict):
    """Generate a strategy variant for each specified token."""
    from app.services.strategy_builder import generate_token_sweep

    pine_code = data.get("pine_code", "")
    tokens = data.get("tokens", [])
    strategy_name = data.get("strategy_name", "Strategy")

    if not pine_code or not tokens:
        raise HTTPException(status_code=400, detail="Provide pine_code and tokens list")

    variants = generate_token_sweep(pine_code, tokens, strategy_name)
    return {"variants": variants, "total": len(variants)}


@router.get("/backtests/compare")
async def compare_strategies():
    """Compare all backtests across strategies and tokens."""
    from app.services.strategy_builder import compare_strategies

    backtests = get_backtests(limit=500)
    comparison = compare_strategies(backtests)
    return comparison


@router.get("/backtests/templates")
async def list_templates():
    """List available strategy templates."""
    from app.services.strategy_builder import get_templates
    return {"templates": get_templates()}


@router.get("/backtests/templates/{template_id}")
async def get_template(template_id: str):
    """Get the full Pine code for a strategy template."""
    from app.services.strategy_builder import get_template_code, TEMPLATES

    if template_id not in TEMPLATES:
        raise HTTPException(status_code=404, detail=f"Template '{template_id}' not found")

    meta = TEMPLATES[template_id]
    return {
        "id": template_id,
        "name": meta["name"],
        "description": meta["description"],
        "pine_code": meta["pine_code"],
    }


def _parse_xlsx(filepath):
    """Parse a TradingView xlsx export into a backtest dict."""
    import re
    try:
        import openpyxl
    except ImportError:
        return None

    wb = openpyxl.load_workbook(filepath, data_only=True)
    filename = filepath.name

    def sheet_dict(name):
        if name not in wb.sheetnames:
            return {}
        ws = wb[name]
        d = {}
        for row in ws.iter_rows(max_col=6, values_only=False):
            key = row[0].value
            if key and key not in d:
                d[key] = {c.column_letter: c.value for c in row[1:] if c.value is not None}
        return d

    perf = sheet_dict("Performance")
    trades = sheet_dict("Trades analysis")
    risk = sheet_dict("Risk-adjusted performance")
    props = sheet_dict("Properties")

    def val(d, key, col="B"):
        return d.get(key, {}).get(col)

    def pct(d, key):
        return d.get(key, {}).get("C")

    # Strategy name/version from filename
    name_match = re.match(r"(.+?)_(v[\d.]+)", filename)
    if name_match:
        strategy_name = name_match.group(1).replace("_", " ")
        version = name_match.group(2)
    else:
        strategy_name = filename.split("_")[0]
        version = "unknown"

    # Symbol from Properties sheet
    symbol_raw = val(props, "Symbol", "B") or ""
    symbol = symbol_raw.split(":")[-1] if ":" in symbol_raw else symbol_raw

    # Timeframe
    tf_raw = val(props, "Timeframe", "B") or ""
    tf_map = {"4 hours": "4H", "1 hour": "1H", "2 hours": "2H", "1 day": "D", "1 week": "W",
              "240": "4H", "60": "1H", "120": "2H", "d": "D", "w": "W"}
    timeframe = tf_map.get(str(tf_raw).lower(), tf_raw)

    # Period from Properties
    trading_range = val(props, "Trading range", "B") or val(props, "Backtesting range", "B") or ""
    period_start, period_end = None, None
    if "—" in str(trading_range):
        parts = str(trading_range).split("—")
        period_start = parts[0].strip()
        period_end = parts[1].strip()

    total = int(val(trades, "Total trades", "B") or 0)
    winners = int(val(trades, "Winning trades", "B") or 0)
    losers = int(val(trades, "Losing trades", "B") or 0)

    return {
        "strategy_name": strategy_name,
        "version": version,
        "timeframe": timeframe,
        "symbol": symbol,
        "period_start": period_start,
        "period_end": period_end,
        "initial_capital": val(perf, "Initial capital", "B"),
        "net_profit_usd": val(perf, "Net profit", "B"),
        "net_profit_pct": pct(perf, "Net profit"),
        "gross_profit": val(perf, "Gross profit", "B"),
        "gross_loss": val(perf, "Gross loss", "B"),
        "profit_factor": val(risk, "Profit factor", "B"),
        "total_trades": total,
        "winning_trades": winners,
        "losing_trades": losers,
        "win_rate": pct(trades, "Percent profitable"),
        "avg_win": val(trades, "Avg winning trade", "B"),
        "avg_loss": val(trades, "Avg losing trade", "B"),
        "win_loss_ratio": val(trades, "Ratio avg win / avg loss", "B"),
        "largest_win": val(trades, "Largest winning trade", "B"),
        "largest_loss": val(trades, "Largest losing trade", "B"),
        "max_drawdown": val(perf, "Max equity drawdown (intrabar)", "B"),
        "sharpe_ratio": val(risk, "Sharpe ratio", "B"),
        "sortino_ratio": val(risk, "Sortino ratio", "B"),
        "long_trades": int(val(trades, "Total trades", "D") or 0),
        "long_win_rate": pct(trades, "Percent profitable") if not val(trades, "Total trades", "D") else None,
        "long_pnl": val(perf, "Net profit", "D"),
        "short_trades": 0,
        "short_win_rate": None,
        "short_pnl": None,
        "source_file": filename,
        "notes": "",
        "status": "tested",
        "created_at": _extract_date_from_filename(filename),
    }


def _extract_date_from_filename(filename: str):
    """Extract date from filename like ..._2026-03-31.xlsx → 2026-03-31T00:00:00"""
    import re
    m = re.search(r"(\d{4}-\d{2}-\d{2})", filename)
    if m:
        return m.group(1) + "T00:00:00"
    return None


# ── Kalshi Event Contracts ────────────────────────────────────────

@router.get("/kalshi/status")
async def kalshi_status():
    """Get Kalshi connection status, balance, and stats."""
    from app.services.kalshi_client import get_async_kalshi_client

    client = get_async_kalshi_client()
    stats = get_kalshi_stats()

    result = {
        "enabled": client.enabled,
        "mode": client.mode,
        "configured": bool(client.api_key_id and client.private_key_path),
        "stats": stats,
    }

    if client.enabled and client.api_key_id:
        try:
            balance = await client.get_balance()
            result["balance"] = balance
            result["connected"] = True
        except Exception as e:
            result["connected"] = False
            result["error"] = str(e)[:100]
    else:
        result["connected"] = False

    return result


@router.get("/kalshi/order-health")
async def kalshi_order_health(limit: int = 50):
    """Get order success/failure stats and recent failures."""
    from app.services.kalshi_client import get_async_kalshi_client

    client = get_async_kalshi_client()
    if not client.enabled:
        return {"success_count": 0, "failure_count": 0, "recent_failures": []}
    return client.get_order_health(limit=min(limit, 200))


@router.get("/kalshi/markets")
async def kalshi_markets(query: str = "", status: str = "open", limit: int = 20):
    """Search or list Kalshi markets."""
    from app.services.kalshi_client import get_async_kalshi_client

    client = get_async_kalshi_client()
    if not client.enabled:
        raise HTTPException(400, "Kalshi is not enabled in config")

    try:
        if query:
            markets = await client.search_markets(query, limit=limit)
        else:
            markets = await client.get_markets(status=status, limit=limit)
        return {"markets": markets, "total": len(markets)}
    except Exception as e:
        raise HTTPException(500, f"Kalshi API error: {e}")


@router.get("/kalshi/market/{ticker}")
async def kalshi_market_detail(ticker: str):
    """Get detailed market info including orderbook."""
    from app.services.kalshi_client import get_async_kalshi_client

    client = get_async_kalshi_client()
    if not client.enabled:
        raise HTTPException(400, "Kalshi is not enabled in config")

    try:
        market = await client.get_market(ticker)
        orderbook = await client.get_orderbook(ticker)
        trades = await client.get_market_trades(ticker, limit=10)
        return {"market": market, "orderbook": orderbook, "recent_trades": trades}
    except Exception as e:
        raise HTTPException(500, f"Kalshi API error: {e}")


@router.get("/kalshi/positions")
async def kalshi_positions(status: str = "all"):
    """Get Kalshi positions from local DB + live data."""
    from app.services.kalshi_client import get_async_kalshi_client

    db_positions = get_kalshi_positions(status=status)

    client = get_async_kalshi_client()
    live_positions = []
    if client.enabled and client.api_key_id:
        try:
            live_positions = await client.get_positions()
        except Exception:
            pass

    return {
        "db_positions": db_positions,
        "live_positions": live_positions,
        "stats": get_kalshi_stats(),
    }


@router.get("/kalshi/trades")
async def kalshi_trade_history(limit: int = 50):
    """Get Kalshi trade history from local DB."""
    trades = get_kalshi_trades(limit=min(limit, 200))
    return {"trades": trades, "total": len(trades)}


@router.post("/kalshi/order")
async def kalshi_place_order(body: dict):
    """Place a Kalshi order. Body: {ticker, side, price, count?, action?}"""
    from app.services.kalshi_client import get_async_kalshi_client
    from app.database import insert_kalshi_trade

    client = get_async_kalshi_client()
    if not client.enabled:
        raise HTTPException(400, "Kalshi is not enabled in config")

    ticker = body.get("ticker")
    side = body.get("side")
    price = body.get("price")
    if not all([ticker, side, price]):
        raise HTTPException(400, "Required: ticker, side (yes/no), price (cents 1-99)")

    action = body.get("action", "buy")
    count = body.get("count")

    try:
        result = await client.place_order(
            ticker=ticker,
            side=side,
            action=action,
            count=count,
            yes_price=price if side == "yes" else None,
            no_price=price if side == "no" else None,
        )

        # Log to DB
        insert_kalshi_trade({
            "order_id": result.get("order", {}).get("order_id", ""),
            "ticker": ticker,
            "title": body.get("title", ""),
            "side": side,
            "action": action,
            "count": count or client.default_count,
            "price_cents": price,
            "total_cost_cents": price * (count or client.default_count),
            "status": result.get("order", {}).get("status", "pending"),
        })

        return {"status": "ok", "result": result}
    except Exception as e:
        raise HTTPException(500, f"Order failed: {e}")


@router.get("/kalshi/portfolio")
async def kalshi_portfolio():
    """Get full Kalshi portfolio summary with live P&L."""
    from app.services.kalshi_client import get_async_kalshi_client

    client = get_async_kalshi_client()
    if not client.enabled:
        raise HTTPException(400, "Kalshi is not enabled in config")

    try:
        summary = await client.get_portfolio_summary()
        summary["db_stats"] = get_kalshi_stats()
        return summary
    except Exception as e:
        raise HTTPException(500, f"Portfolio error: {e}")


@router.get("/kalshi/unified-pnl")
async def kalshi_unified_pnl():
    """Aggregated P&L across all Kalshi bots."""
    from app.services.kalshi_market_maker import get_market_maker
    from app.services.kalshi_spread_bot import get_spread_bot
    from app.services.kalshi_technical_bot import get_technical_bot
    from app.services.kalshi_ai_agent import get_ai_agent_bot

    bots = {}
    total_pnl_cents = 0

    # Market Maker
    try:
        mm = get_market_maker()
        mm_status = mm.get_status()
        mm_pnl = mm._total_pnl_cents
        bots["market_maker"] = {
            "running": mm_status.get("running", False),
            "pnl_cents": mm_pnl,
            "pnl_usd": round(mm_pnl / 100, 2),
            "markets": mm_status.get("active_markets", 0),
            "fills": mm_status.get("total_fills", 0),
        }
        total_pnl_cents += mm_pnl
    except Exception:
        bots["market_maker"] = {"running": False, "pnl_cents": 0, "pnl_usd": 0.0}

    # Spread Bot
    try:
        sb = get_spread_bot()
        sb_status = sb.get_status()
        sb_pnl = sb._total_pnl_cents
        bots["spread_bot"] = {
            "running": sb_status.get("running", False),
            "pnl_cents": sb_pnl,
            "pnl_usd": round(sb_pnl / 100, 2),
            "markets": sb_status.get("active_markets", 0),
        }
        total_pnl_cents += sb_pnl
    except Exception:
        bots["spread_bot"] = {"running": False, "pnl_cents": 0, "pnl_usd": 0.0}

    # Technical Bot
    try:
        tb = get_technical_bot()
        tb_status = tb.get_status()
        tb_trades = tb_status.get("trades_executed", 0)
        bots["technical"] = {
            "running": tb_status.get("running", False),
            "trades": tb_trades,
            "signals": tb_status.get("signals_generated", 0),
        }
    except Exception:
        bots["technical"] = {"running": False, "trades": 0}

    # AI Agent
    try:
        ai = get_ai_agent_bot()
        ai_status = ai.get_status()
        ai_trades = ai_status.get("trades_executed", 0)
        bots["ai_agent"] = {
            "running": ai_status.get("running", False),
            "trades": ai_trades,
            "analyses": ai_status.get("analyses_completed", 0),
        }
    except Exception:
        bots["ai_agent"] = {"running": False, "trades": 0}

    # DB stats
    db_stats = get_kalshi_stats()

    return {
        "total_pnl_cents": total_pnl_cents,
        "total_pnl_usd": round(total_pnl_cents / 100, 2),
        "bots": bots,
        "db_stats": db_stats,
    }


# ── Kalshi Arbitrage ──────────────────────────────────────────────

@router.get("/kalshi/arbitrage/status")
async def kalshi_arb_status():
    """Get arbitrage scanner status."""
    from app.services.kalshi_arbitrage import get_arbitrage_scanner
    scanner = get_arbitrage_scanner()
    return scanner.get_status()


@router.get("/kalshi/arbitrage/opportunities")
async def kalshi_arb_opportunities():
    """Get current arbitrage opportunities."""
    from app.services.kalshi_arbitrage import get_arbitrage_scanner
    scanner = get_arbitrage_scanner()
    return {"opportunities": scanner.get_opportunities()}


@router.post("/kalshi/arbitrage/scan")
async def kalshi_arb_scan_now():
    """Manually trigger an arbitrage scan."""
    from app.services.kalshi_arbitrage import get_arbitrage_scanner
    scanner = get_arbitrage_scanner()
    results = await scanner.scan_all()
    return {"opportunities": results, "total": len(results)}


@router.post("/kalshi/arbitrage/toggle")
async def kalshi_arb_toggle(body: dict):
    """Toggle auto-execute or scanner. Body: {auto_execute?: bool, enabled?: bool}"""
    from app.services.kalshi_arbitrage import get_arbitrage_scanner
    scanner = get_arbitrage_scanner()

    if "auto_execute" in body:
        scanner.auto_execute = bool(body["auto_execute"])
    if "enabled" in body:
        if body["enabled"] and not scanner._running:
            scanner.enabled = True
            scanner.start()
        elif not body["enabled"]:
            scanner.stop()
            scanner.enabled = False

    return scanner.get_status()


# ── Kalshi Spread Bot ─────────────────────────────────────────────

@router.get("/kalshi/spread/status")
async def kalshi_spread_status():
    """Get spread bot status and all market states."""
    from app.services.kalshi_spread_bot import get_spread_bot
    bot = get_spread_bot()
    return bot.get_status()


@router.post("/kalshi/spread/start")
async def kalshi_spread_start():
    """Start the spread bot."""
    from app.services.kalshi_spread_bot import get_spread_bot
    bot = get_spread_bot()
    if bot._running:
        return {"status": "already_running"}
    bot.start()
    return {"status": "started"}


@router.post("/kalshi/spread/stop")
async def kalshi_spread_stop():
    """Stop the spread bot and cancel all orders."""
    from app.services.kalshi_spread_bot import get_spread_bot
    bot = get_spread_bot()
    await bot.stop()
    return {"status": "stopped"}


@router.post("/kalshi/spread/kill")
async def kalshi_spread_kill():
    """Emergency kill switch."""
    from app.services.kalshi_spread_bot import get_spread_bot
    bot = get_spread_bot()
    bot.kill()
    return {"status": "killed"}


@router.post("/kalshi/spread/flatten")
async def kalshi_spread_flatten(body: dict = {}):
    """Flatten positions. Body: {ticker: "..."} for one market, or {} for all."""
    from app.services.kalshi_spread_bot import get_spread_bot
    bot = get_spread_bot()
    ticker = body.get("ticker")
    if ticker:
        await bot.flatten_market(ticker)
        return {"status": "flattened", "ticker": ticker}
    else:
        await bot.flatten_all()
        return {"status": "flattened_all"}


@router.post("/kalshi/spread/market")
async def kalshi_spread_add_market(body: dict):
    """Add/remove a market from spread bot. Body: {ticker, action: "add"|"remove"}"""
    from app.services.kalshi_spread_bot import get_spread_bot
    bot = get_spread_bot()
    ticker = body.get("ticker")
    action = body.get("action", "add")
    if not ticker:
        raise HTTPException(400, "ticker required")
    if action == "add":
        bot.add_market(ticker)
    elif action == "remove":
        bot.remove_market(ticker)
        await bot.flatten_market(ticker)
    return {"status": "ok", "targets": bot.target_tickers}


# ── Kalshi Whale Tracker ──────────────────────────────────────────

@router.get("/kalshi/whales")
async def kalshi_whales(limit: int = 50):
    """Get detected whale trades."""
    from app.services.kalshi_whale_tracker import get_whale_tracker
    tracker = get_whale_tracker()
    return {"whales": tracker.get_whales(limit=min(limit, 100)), "status": tracker.get_status()}


@router.post("/kalshi/whales/scan")
async def kalshi_whale_scan():
    """Manually trigger a whale scan."""
    from app.services.kalshi_whale_tracker import get_whale_tracker
    tracker = get_whale_tracker()
    results = await tracker.scan()
    return {"new_whales": results, "total": len(results)}


@router.post("/kalshi/whales/toggle")
async def kalshi_whale_toggle(body: dict):
    """Start/stop whale tracker. Body: {enabled: bool}"""
    from app.services.kalshi_whale_tracker import get_whale_tracker
    tracker = get_whale_tracker()
    if body.get("enabled") and not tracker._running:
        tracker.enabled = True
        tracker.start()
    elif not body.get("enabled"):
        tracker.stop()
        tracker.enabled = False
    return tracker.get_status()


# ── Kalshi Technical Bot (MACD/CCI) ──────────────────────────────

@router.get("/kalshi/tech/status")
async def kalshi_tech_status():
    """Get technical bot status."""
    from app.services.kalshi_technical_bot import get_technical_bot
    bot = get_technical_bot()
    return bot.get_status()


@router.get("/kalshi/tech/signals")
async def kalshi_tech_signals(limit: int = 50):
    """Get recent technical signals."""
    from app.services.kalshi_technical_bot import get_technical_bot
    bot = get_technical_bot()
    return {"signals": bot.get_signals(limit=min(limit, 100)), "status": bot.get_status()}


@router.post("/kalshi/tech/scan")
async def kalshi_tech_scan():
    """Manually trigger a technical analysis scan."""
    from app.services.kalshi_technical_bot import get_technical_bot
    bot = get_technical_bot()
    results = await bot.scan_all()
    return {"signals": results, "total": len(results)}


@router.post("/kalshi/tech/toggle")
async def kalshi_tech_toggle(body: dict):
    """Start/stop technical bot. Body: {enabled?: bool, auto_trade?: bool}"""
    from app.services.kalshi_technical_bot import get_technical_bot
    bot = get_technical_bot()
    if "auto_trade" in body:
        bot.auto_trade = bool(body["auto_trade"])
    if "enabled" in body:
        if body["enabled"] and not bot._running:
            bot.enabled = True
            bot.start()
        elif not body["enabled"]:
            bot.stop()
            bot.enabled = False
    return bot.get_status()


# ── Kalshi AI Agent Bot ──────────────────────────────────────────

@router.get("/kalshi/ai/status")
async def kalshi_ai_status():
    """Get AI agent bot status."""
    from app.services.kalshi_ai_agent import get_ai_agent_bot
    bot = get_ai_agent_bot()
    return bot.get_status()


@router.get("/kalshi/ai/decisions")
async def kalshi_ai_decisions(limit: int = 20):
    """Get recent AI consensus decisions."""
    from app.services.kalshi_ai_agent import get_ai_agent_bot
    bot = get_ai_agent_bot()
    return {"decisions": bot.get_decisions(limit=min(limit, 50)), "status": bot.get_status()}


@router.post("/kalshi/ai/analyze")
async def kalshi_ai_analyze():
    """Manually trigger AI agent analysis."""
    from app.services.kalshi_ai_agent import get_ai_agent_bot
    bot = get_ai_agent_bot()
    results = await bot.analyze_markets()
    return {"decisions": results, "total": len(results)}


@router.post("/kalshi/ai/toggle")
async def kalshi_ai_toggle(body: dict):
    """Start/stop AI agent bot. Body: {enabled?, auto_trade?}"""
    from app.services.kalshi_ai_agent import get_ai_agent_bot
    bot = get_ai_agent_bot()
    if "auto_trade" in body:
        bot.auto_trade = bool(body["auto_trade"])
    if "enabled" in body:
        if body["enabled"] and not bot._running:
            bot.enabled = True
            bot.start()
        elif not body["enabled"]:
            bot.stop()
            bot.enabled = False
    return bot.get_status()


# =========================================================================
# KALSHI SPORTS SCANNER
# =========================================================================

@router.get("/kalshi/sports/status")
async def kalshi_sports_status():
    """Get sports scanner status."""
    from app.services.kalshi_sports_scanner import get_sports_scanner
    return get_sports_scanner().get_status()


@router.get("/kalshi/sports/markets")
async def kalshi_sports_markets(league: str = None, limit: int = 50):
    """Get tracked sports markets, optionally by league."""
    from app.services.kalshi_sports_scanner import get_sports_scanner
    return {"markets": get_sports_scanner().get_markets_by_league(league), "status": get_sports_scanner().get_status()}


@router.get("/kalshi/sports/value-bets")
async def kalshi_sports_value_bets(limit: int = 20):
    """Get recent value bet opportunities."""
    from app.services.kalshi_sports_scanner import get_sports_scanner
    return {"value_bets": get_sports_scanner().get_value_bets(limit=min(limit, 50))}


@router.post("/kalshi/sports/scan")
async def kalshi_sports_scan():
    """Manually trigger sports market scan."""
    from app.services.kalshi_sports_scanner import get_sports_scanner
    result = await get_sports_scanner().scan()
    return result


@router.post("/kalshi/sports/toggle")
async def kalshi_sports_toggle(body: dict):
    """Start/stop sports scanner. Body: {enabled?, auto_trade?}"""
    from app.services.kalshi_sports_scanner import get_sports_scanner
    scanner = get_sports_scanner()
    if "auto_trade" in body:
        scanner.auto_trade = bool(body["auto_trade"])
    if "enabled" in body:
        if body["enabled"] and not scanner._running:
            scanner.enabled = True
            scanner.start()
        elif not body["enabled"]:
            scanner.stop()
            scanner.enabled = False
    return scanner.get_status()


# =========================================================================
# KALSHI MARKET MAKER FRAMEWORK
# =========================================================================

@router.get("/kalshi/mm/status")
async def kalshi_mm_status():
    """Get market maker status."""
    from app.services.kalshi_market_maker import get_market_maker
    return get_market_maker().get_status()


@router.get("/kalshi/mm/fills")
async def kalshi_mm_fills(limit: int = 50):
    """Get recent fills."""
    from app.services.kalshi_market_maker import get_market_maker
    return {"fills": get_market_maker().get_fills(limit=min(limit, 200))}


@router.get("/kalshi/mm/pnl")
async def kalshi_mm_pnl():
    """Get P&L breakdown by market and strategy."""
    from app.services.kalshi_market_maker import get_market_maker
    mm = get_market_maker()
    return {
        "by_market": mm.get_pnl_by_market(),
        "by_strategy": mm.get_pnl_by_strategy(),
        "total_pnl_cents": mm._total_pnl_cents,
        "total_pnl_usd": round(mm._total_pnl_cents / 100, 2),
    }


@router.post("/kalshi/mm/start")
async def kalshi_mm_start():
    """Start the market maker."""
    from app.services.kalshi_market_maker import get_market_maker
    mm = get_market_maker()
    mm.enabled = True
    mm.start()
    return mm.get_status()


@router.post("/kalshi/mm/stop")
async def kalshi_mm_stop():
    """Stop the market maker gracefully."""
    from app.services.kalshi_market_maker import get_market_maker
    mm = get_market_maker()
    await mm.stop()
    mm.enabled = False
    return mm.get_status()


@router.post("/kalshi/mm/kill")
async def kalshi_mm_kill():
    """Emergency kill switch."""
    from app.services.kalshi_market_maker import get_market_maker
    mm = get_market_maker()
    mm.kill()
    return mm.get_status()


@router.post("/kalshi/mm/flatten")
async def kalshi_mm_flatten(body: dict = None):
    """Flatten one or all markets. Body: {ticker?}"""
    from app.services.kalshi_market_maker import get_market_maker
    mm = get_market_maker()
    if body and body.get("ticker"):
        result = await mm.flatten_market(body["ticker"])
        return result
    await mm.flatten_all()
    return {"status": "all_flattened"}


# =========================================================================
# KALSHI ESPORTS SCANNER
# =========================================================================

@router.get("/kalshi/esports/status")
async def kalshi_esports_status():
    """Get esports scanner status."""
    from app.services.kalshi_esports_scanner import get_esports_scanner
    return get_esports_scanner().get_status()


@router.get("/kalshi/esports/markets")
async def kalshi_esports_markets(game: str = None, limit: int = 50):
    """Get tracked esports markets, optionally by game."""
    from app.services.kalshi_esports_scanner import get_esports_scanner
    return {"markets": get_esports_scanner().get_markets_by_game(game), "status": get_esports_scanner().get_status()}


@router.get("/kalshi/esports/value-bets")
async def kalshi_esports_value_bets(limit: int = 20):
    """Get recent esports value bet opportunities."""
    from app.services.kalshi_esports_scanner import get_esports_scanner
    return {"value_bets": get_esports_scanner().get_value_bets(limit=min(limit, 50))}


@router.post("/kalshi/esports/scan")
async def kalshi_esports_scan():
    """Manually trigger esports market scan."""
    from app.services.kalshi_esports_scanner import get_esports_scanner
    result = await get_esports_scanner().scan()
    return result


@router.post("/kalshi/esports/toggle")
async def kalshi_esports_toggle(body: dict):
    """Start/stop esports scanner. Body: {enabled?, auto_trade?}"""
    from app.services.kalshi_esports_scanner import get_esports_scanner
    scanner = get_esports_scanner()
    if "auto_trade" in body:
        scanner.auto_trade = bool(body["auto_trade"])
    if "enabled" in body:
        if body["enabled"] and not scanner._running:
            scanner.enabled = True
            scanner.start()
        elif not body["enabled"]:
            scanner.stop()
            scanner.enabled = False
    return scanner.get_status()


# ── Kalshi WebSocket Feed ────────────────────────────────────────

@router.get("/kalshi/ws/status")
async def kalshi_ws_status():
    """Get WebSocket feed status."""
    from app.services.kalshi_ws_feed import get_kalshi_ws_feed
    return get_kalshi_ws_feed().get_status()


@router.get("/kalshi/ws/trades")
async def kalshi_ws_trades(limit: int = 50):
    """Get live trade feed from WebSocket."""
    from app.services.kalshi_ws_feed import get_kalshi_ws_feed
    return {"trades": get_kalshi_ws_feed().get_trade_log(limit=min(limit, 200))}


@router.get("/kalshi/ws/orderbook/{ticker}")
async def kalshi_ws_orderbook(ticker: str):
    """Get live orderbook for a subscribed ticker."""
    from app.services.kalshi_ws_feed import get_kalshi_ws_feed
    book = get_kalshi_ws_feed().get_orderbook(ticker)
    if not book:
        raise HTTPException(404, f"No live orderbook for {ticker}")
    return book.to_dict()


@router.post("/kalshi/ws/subscribe")
async def kalshi_ws_subscribe(body: dict):
    """Subscribe to a market ticker. Body: {ticker: "..."}"""
    from app.services.kalshi_ws_feed import get_kalshi_ws_feed
    ticker = body.get("ticker", "")
    if not ticker:
        raise HTTPException(400, "ticker required")
    await get_kalshi_ws_feed().subscribe(ticker)
    return {"subscribed": ticker}


# ── Kalshi Risk Manager (Circuit Breaker) ────────────────────────

@router.get("/kalshi/risk/status")
async def kalshi_risk_status():
    """Get circuit breaker status."""
    from app.services.kalshi_risk_manager import get_risk_manager
    return get_risk_manager().get_status()


@router.post("/kalshi/risk/reset")
async def kalshi_risk_reset():
    """Manually reset the circuit breaker to resume trading."""
    from app.services.kalshi_risk_manager import get_risk_manager
    return get_risk_manager().reset()


@router.post("/kalshi/risk/reconcile")
async def kalshi_risk_reconcile():
    """Force an immediate reconcile of category exposure from live Kalshi positions.

    Use when the in-memory `_category_exposure` counter has drifted from reality
    (e.g., resting orders that never filled left phantom exposure on the books).
    The periodic `_check` loop also reconciles every ~30s automatically.
    """
    from app.services.kalshi_risk_manager import get_risk_manager
    rm = get_risk_manager()
    before = {k: v for k, v in rm._category_exposure.items()}
    await rm._reconcile_category_exposure()
    after = {k: v for k, v in rm._category_exposure.items()}
    return {
        "reconciled": True,
        "before_cents": before,
        "after_cents": after,
    }


def _parse_report_txt(text: str, filename: str) -> dict:
    """Parse a .report.txt file into a backtest dict."""
    import re

    def extract(pattern, txt, group=1, as_float=True):
        m = re.search(pattern, txt, re.MULTILINE)
        if not m:
            return None
        val = m.group(group).strip().replace(",", "").replace("$", "").replace("+", "").replace("%", "")
        if as_float:
            try:
                return float(val)
            except ValueError:
                return None
        return val

    # Parse strategy name and version from filename
    # e.g. SOL_Confluence_Pro_v3.5_COINBASE_SOLUSD_2026-03-29.report.txt
    # or SOL_Mean_Reversion_v1.3_—_4H_Backtest_BINANCE_SOLUSDT_2026-03-30.report.txt
    name_match = re.match(r"(.+?)_(v[\d.]+)", filename)
    if name_match:
        strategy_name = name_match.group(1).replace("_", " ")
        version = name_match.group(2)
    else:
        strategy_name = filename.split("_")[0]
        version = "unknown"

    symbol = extract(r"Symbol\s*:\s*(\S+)", text, as_float=False) or ""
    tf_raw = extract(r"Timeframe:\s*(.+?)$", text, as_float=False) or ""
    tf_raw = tf_raw.strip()
    # Normalize: "4 hours" → "4H", "1 hour" → "1H", "D" → "D"
    tf_map = {"4 hours": "4H", "1 hour": "1H", "2 hours": "2H", "1 day": "D", "1 week": "W",
              "240": "4H", "60": "1H", "120": "2H", "d": "D", "w": "W"}
    timeframe = tf_map.get(tf_raw.lower(), tf_raw)
    period_line = re.search(r"Period\s*:\s*(.+?)$", text, re.MULTILINE)
    period_start, period_end = None, None
    if period_line:
        parts = period_line.group(1).split("—")
        if len(parts) == 2:
            period_start = parts[0].strip()
            period_end = parts[1].strip()

    capital = extract(r"Capital\s*:\s*\$([\d,.]+)", text)

    # Determine status from filename/strategy
    status = "tested"

    return {
        "strategy_name": strategy_name,
        "version": version,
        "timeframe": timeframe,
        "symbol": symbol,
        "period_start": period_start,
        "period_end": period_end,
        "initial_capital": capital,
        "net_profit_usd": extract(r"Net Profit\s*:\s*\$([+\-\d,.]+)", text),
        "net_profit_pct": extract(r"Net Profit\s*:.*?\(([+\-\d,.]+)%\)", text),
        "gross_profit": extract(r"Gross Profit\s*:\s*\$([\d,.]+)", text),
        "gross_loss": extract(r"Gross Loss\s*:\s*-?\$?([\d,.]+)", text),
        "profit_factor": extract(r"Profit Factor\s*:\s*([\d.]+)", text),
        "total_trades": int(extract(r"Total Trades\s*:\s*(\d+)", text) or 0),
        "winning_trades": int(extract(r"Winners.*?:\s*(\d+)", text) or 0),
        "losing_trades": int(extract(r"Losers.*?:\s*(\d+)", text) or extract(r"Winners\s*/\s*Losers\s*:\s*\d+\s*/\s*(\d+)", text) or 0),
        "win_rate": extract(r"Win Rate\s*:\s*([\d.]+)", text),
        "avg_win": extract(r"Avg Win\s*:\s*\$([+\-\d,.]+)", text),
        "avg_loss": extract(r"Avg Loss\s*:\s*-?\$?([\d,.]+)", text),
        "win_loss_ratio": extract(r"Win/Loss Ratio\s*:\s*([\d.]+)", text),
        "largest_win": extract(r"Largest Win\s*:\s*\$([+\-\d,.]+)", text),
        "largest_loss": extract(r"Largest Loss\s*:\s*-?\$?([\d,.]+)", text),
        "max_drawdown": extract(r"Max Drawdown \(EOB\)\s*:\s*-?\$?([\d,.]+)", text),
        "sharpe_ratio": extract(r"Sharpe Ratio\s*:\s*([+\-\d.]+)", text),
        "sortino_ratio": extract(r"Sortino Ratio\s*:\s*([+\-\d.]+)", text),
        "long_trades": int(extract(r"Long\s*:\s*(\d+)\s*trades", text) or 0),
        "long_win_rate": extract(r"Long\s*:.*?\(([\d.]+)%\)", text),
        "long_pnl": extract(r"Long\s*:.*?P&L:\s*\$([+\-\d,.]+)", text),
        "short_trades": int(extract(r"Short\s*:\s*(\d+)\s*trades", text) or 0),
        "short_win_rate": extract(r"Short\s*:.*?\(([\d.]+)%\)", text),
        "short_pnl": extract(r"Short\s*:.*?P&L:\s*\$([+\-\d,.]+)", text),
        "source_file": filename,
        "notes": "",
        "status": status,
        "created_at": _extract_date_from_filename(filename),
    }


# =============================================================================
# PAPER TRADING ENDPOINTS
# =============================================================================


@router.get("/paper/portfolio")
async def get_paper_portfolio():
    """Get current paper trading portfolio state."""
    paper = get_paper_trader()
    return paper.get_paper_portfolio()


@router.get("/paper/trades")
async def get_paper_trade_history(limit: int = 50):
    """Get paper trade history."""
    paper = get_paper_trader()
    trades = paper.get_paper_trades(limit=min(limit, 200))
    return {"trades": trades, "total": len(trades)}


@router.get("/paper/stats")
async def get_paper_trading_stats():
    """Get paper trading statistics."""
    paper = get_paper_trader()
    return paper.get_paper_stats()


@router.post("/paper/reset")
async def reset_paper_portfolio():
    """Reset paper portfolio to starting balance."""
    paper = get_paper_trader()
    return paper.reset()


@router.post("/paper/toggle")
async def toggle_paper_mode():
    """Enable/disable paper trading mode."""
    import yaml

    config_path = Path("config.yaml")
    if not config_path.exists():
        raise HTTPException(status_code=500, detail="config.yaml not found")

    with open(config_path) as f:
        cfg = yaml.safe_load(f)

    current = cfg.get("paper_trading", {}).get("enabled", False)
    cfg.setdefault("paper_trading", {})["enabled"] = not current

    with open(config_path, "w") as f:
        yaml.dump(cfg, f, default_flow_style=False)

    reload_config()
    new_state = not current
    logger.info(f"Paper trading mode {'enabled' if new_state else 'disabled'}")
    return {"enabled": new_state, "status": "ok"}


# =============================================================================
# PORTFOLIO REBALANCER
# =============================================================================

@router.get("/rebalancer/status")
async def rebalancer_status():
    """Current allocations vs targets, drift, last rebalance."""
    from app.services.portfolio_rebalancer import get_rebalancer
    rebalancer = get_rebalancer()
    return await rebalancer.get_status()


@router.post("/rebalancer/calculate")
async def rebalancer_calculate():
    """Dry-run: show what trades would be needed to rebalance."""
    from app.services.portfolio_rebalancer import get_rebalancer
    rebalancer = get_rebalancer()
    try:
        current = await rebalancer.get_current_allocations()
        trades = rebalancer.calculate_rebalance(current)
        return {
            "status": "calculated",
            "current_allocations": {k: round(v, 2) for k, v in current["allocations"].items()},
            "targets": rebalancer.targets,
            "total_usd": round(current["total_usd"], 2),
            "trades": trades,
        }
    except Exception as e:
        logger.error(f"Rebalancer calculate error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/rebalancer/execute")
async def rebalancer_execute():
    """Execute rebalance now."""
    from app.services.portfolio_rebalancer import get_rebalancer
    rebalancer = get_rebalancer()
    result = await rebalancer.execute_rebalance()
    if result.get("status") == "error":
        raise HTTPException(status_code=500, detail=result.get("error"))
    return result


@router.post("/rebalancer/toggle")
async def rebalancer_toggle():
    """Enable/disable auto-rebalance."""
    from app.services.portfolio_rebalancer import get_rebalancer
    rebalancer = get_rebalancer()
    return rebalancer.toggle_auto()


# ── Prometheus-style Metrics ─────────────────────────────────────

@router.get("/metrics", response_class=PlainTextResponse)
async def prometheus_metrics():
    """Prometheus-compatible metrics endpoint (plaintext)."""
    lines = []

    def gauge(name, value, help_text="", labels=None):
        if help_text:
            lines.append(f"# HELP {name} {help_text}")
            lines.append(f"# TYPE {name} gauge")
        label_str = ""
        if labels:
            label_str = "{" + ",".join(f'{k}="{v}"' for k, v in labels.items()) + "}"
        lines.append(f"{name}{label_str} {value}")

    def counter(name, value, help_text="", labels=None):
        if help_text:
            lines.append(f"# HELP {name} {help_text}")
            lines.append(f"# TYPE {name} counter")
        label_str = ""
        if labels:
            label_str = "{" + ",".join(f'{k}="{v}"' for k, v in labels.items()) + "}"
        lines.append(f"{name}{label_str} {value}")

    # P&L per bot
    try:
        from app.services.kalshi_market_maker import get_market_maker
        mm = get_market_maker()
        gauge("kalshi_pnl_cents", mm._total_pnl_cents,
              "P&L in cents per bot", {"strategy": "market_maker"})
        gauge("kalshi_bot_running", int(mm._running), labels={"bot": "market_maker"})
    except Exception:
        pass

    try:
        from app.services.kalshi_spread_bot import get_spread_bot
        sb = get_spread_bot()
        gauge("kalshi_pnl_cents", sb._total_pnl_cents, labels={"strategy": "spread_bot"})
        gauge("kalshi_bot_running", int(sb._running), labels={"bot": "spread_bot"})
    except Exception:
        pass

    try:
        from app.services.kalshi_technical_bot import get_technical_bot
        tb = get_technical_bot()
        gauge("kalshi_bot_running", int(tb._running), labels={"bot": "technical"})
        tb_status = tb.get_status()
        counter("kalshi_trades_total", tb_status.get("trades_executed", 0),
                "Total trades placed per bot", {"bot": "technical"})
        counter("kalshi_signals_total", tb_status.get("signals_generated", 0),
                labels={"bot": "technical"})
    except Exception:
        pass

    try:
        from app.services.kalshi_ai_agent import get_ai_agent_bot
        ai = get_ai_agent_bot()
        gauge("kalshi_bot_running", int(ai._running), labels={"bot": "ai_agent"})
        ai_status = ai.get_status()
        counter("kalshi_trades_total", ai_status.get("trades_executed", 0),
                labels={"bot": "ai_agent"})
        counter("kalshi_analyses_total", ai_status.get("analyses_completed", 0),
                "Total AI analyses run", labels={"bot": "ai_agent"})
    except Exception:
        pass

    try:
        from app.services.kalshi_arbitrage import get_arbitrage_scanner
        arb = get_arbitrage_scanner()
        gauge("kalshi_bot_running", int(arb._running), labels={"bot": "arb_scanner"})
    except Exception:
        pass

    try:
        from app.services.kalshi_sports_scanner import get_sports_scanner
        sports = get_sports_scanner()
        gauge("kalshi_bot_running", int(sports._running), labels={"bot": "sports"})
    except Exception:
        pass

    try:
        from app.services.kalshi_esports_scanner import get_esports_scanner
        esports = get_esports_scanner()
        gauge("kalshi_bot_running", int(esports._running), labels={"bot": "esports"})
    except Exception:
        pass

    # Circuit breaker / risk manager
    try:
        from app.services.kalshi_risk_manager import get_risk_manager
        rm = get_risk_manager()
        status = rm.get_status()
        gauge("kalshi_circuit_breaker_tripped", int(status["tripped"]),
              "Whether the circuit breaker is tripped")
        gauge("kalshi_aggregate_pnl_cents", status["current_pnl_cents"],
              "Aggregate P&L across all bots in cents")
        gauge("kalshi_max_daily_loss_cents", status["max_daily_loss_cents"],
              "Daily loss limit in cents")

        # Category exposure
        for cat, cat_data in status.get("categories", {}).items():
            gauge("kalshi_category_used_cents", cat_data["used_cents"],
                  "Category exposure used in cents", {"category": cat})
            gauge("kalshi_category_limit_cents", cat_data["limit_cents"],
                  labels={"category": cat})
    except Exception:
        pass

    # Rate limiter
    try:
        from app.services.kalshi_client import get_kalshi_client
        client = get_kalshi_client()
        gauge("kalshi_rate_limit_tokens", client._limiter._tokens,
              "Rate limit tokens remaining")
    except Exception:
        pass

    # Active positions from DB
    try:
        db_stats = get_kalshi_stats()
        gauge("kalshi_active_positions", db_stats.get("open_positions", 0),
              "Number of open Kalshi positions")
        counter("kalshi_db_trades_total", db_stats.get("total_positions", 0),
                "Total trades recorded in DB")
    except Exception:
        pass

    # WebSocket feed
    try:
        from app.services.kalshi_ws_feed import get_kalshi_ws_feed
        ws = get_kalshi_ws_feed()
        gauge("kalshi_ws_connected", int(ws._connected),
              "Whether Kalshi WebSocket is connected")
        gauge("kalshi_ws_subscriptions", len(ws._subscribed_tickers),
              "Number of WS ticker subscriptions")
    except Exception:
        pass

    lines.append("")
    return "\n".join(lines)


# ── DRY-RUN SIMULATION ENDPOINTS ────────────────────────────────────────────


@router.get("/dryrun/alerts")
async def get_dryrun_alerts():
    """Get all alerts with their live/dry_run mode."""
    from app.services.dry_run_manager import get_dry_run_manager
    drm = get_dry_run_manager()
    return {
        "default_mode": drm.get_default_mode(),
        "alerts": drm.get_all_alerts_with_modes(),
    }


@router.post("/dryrun/toggle")
async def toggle_dryrun_mode(payload: dict):
    """Toggle an alert between live and dry_run mode.

    Body: {"strategy": "BB Squeeze v1.0", "token": "SOL", "timeframe": "4H", "mode": "dry_run"}
    """
    from app.services.dry_run_manager import get_dry_run_manager
    drm = get_dry_run_manager()
    strategy = payload.get("strategy", "")
    token = payload.get("token", "")
    timeframe = payload.get("timeframe", "")
    mode = payload.get("mode", "")
    if not strategy or not token or not mode:
        raise HTTPException(status_code=400, detail="strategy, token, and mode are required")
    return drm.set_mode(strategy, token, timeframe, mode)


@router.get("/dryrun/status")
async def get_dryrun_status():
    """Get dry-run simulation status and recent trades."""
    from app.services.dry_run_manager import get_dry_run_manager
    drm = get_dry_run_manager()
    return drm.get_status()


@router.get("/dryrun/trades")
async def get_dryrun_trades(limit: int = 50):
    """Get simulated trade log."""
    from app.services.dry_run_manager import get_dry_run_manager
    drm = get_dry_run_manager()
    return {"trades": drm.get_sim_trades(limit)}


@router.post("/dryrun/reset")
async def reset_dryrun():
    """Clear all simulated trade history."""
    from app.services.dry_run_manager import get_dry_run_manager
    drm = get_dry_run_manager()
    drm.reset_sim_trades()
    return {"status": "ok", "message": "Simulation history cleared"}


@router.get("/dryrun/summary")
async def get_dryrun_hourly_summary():
    """Get hourly signal summary."""
    from app.services.dry_run_manager import get_dry_run_manager
    drm = get_dry_run_manager()
    return drm.get_hourly_summary()


# ── CONFLUENCE FILTER ENDPOINTS ─────────────────────────────────────────────


@router.get("/confluence/status")
async def get_confluence_status():
    """Get confluence filter status, stats, and pending signals."""
    from app.services.confluence_filter import get_confluence_filter
    cf = get_confluence_filter()
    return cf.get_status()


# ── SOL RISK MANAGER ENDPOINTS ──────────────────────────────────────────────


@router.get("/sol-risk/status")
async def get_sol_risk_status():
    """Get SOL risk manager status — daily P&L, exposure, circuit breaker."""
    from app.services.sol_risk_manager import get_sol_risk_manager
    rm = get_sol_risk_manager()
    return rm.get_status()


@router.post("/sol-risk/reset")
async def reset_sol_risk():
    """Reset SOL risk manager — clear circuit breaker and exposure tracking."""
    from app.services.sol_risk_manager import get_sol_risk_manager
    rm = get_sol_risk_manager()
    rm.reset()
    return {"status": "ok", "message": "SOL risk manager reset"}


# ── Backtest Scorer ─────────────────────────────────────────────────────────


@router.get("/backtest-scorer/experiments")
async def list_backtest_experiments(status: str = None):
    """List all strategy comparison experiments."""
    from app.services.backtest_scorer import get_backtest_scorer
    scorer = get_backtest_scorer()
    return scorer.list_experiments(status=status)


@router.get("/backtest-scorer/experiments/{name}")
async def get_backtest_experiment(name: str):
    """Get full detail for an experiment including trades."""
    from app.services.backtest_scorer import get_backtest_scorer
    scorer = get_backtest_scorer()
    exp = scorer.get_experiment_detail(name)
    if not exp:
        raise HTTPException(status_code=404, detail=f"Experiment '{name}' not found")
    return exp


@router.post("/backtest-scorer/experiments")
async def create_backtest_experiment(body: dict):
    """Create a new strategy comparison experiment.

    Body: {name, baseline, variants: [...], metric?, min_trades?, hypothesis?}
    """
    from app.services.backtest_scorer import get_backtest_scorer
    scorer = get_backtest_scorer()
    try:
        exp = scorer.create_experiment(
            name=body["name"],
            baseline=body["baseline"],
            variants=body["variants"],
            metric=body.get("metric", "pnl_usd"),
            min_trades=body.get("min_trades", 30),
            hypothesis=body.get("hypothesis", ""),
        )
        return {"status": "created", "experiment": exp["name"]}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/backtest-scorer/experiments/{name}/trade")
async def log_backtest_trade(name: str, body: dict):
    """Log a trade to an experiment.

    Body: {variant, metrics: {pnl_usd: ..., win: ..., ...}}
    """
    from app.services.backtest_scorer import get_backtest_scorer
    scorer = get_backtest_scorer()
    try:
        entry = scorer.log_trade(
            experiment_name=name,
            variant=body["variant"],
            metrics=body["metrics"],
            timestamp=body.get("timestamp"),
        )
        return {"status": "logged", "entry": entry}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/backtest-scorer/experiments/{name}/score")
async def score_backtest_experiment(name: str):
    """Score an experiment — run statistical comparison."""
    from app.services.backtest_scorer import get_backtest_scorer
    scorer = get_backtest_scorer()
    try:
        result = scorer.score(name)
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/backtest-scorer/experiments/{name}/import-dry-run")
async def import_dry_run_to_experiment(name: str, body: dict = None):
    """Import trades from dry-run log into an experiment.

    Body (optional): {variant_map: {"STRATEGY NAME": "variant_name", ...}}
    """
    from app.services.backtest_scorer import get_backtest_scorer
    scorer = get_backtest_scorer()
    variant_map = (body or {}).get("variant_map")
    try:
        count = scorer.import_dry_run_trades(name, variant_map=variant_map)
        return {"status": "imported", "trades_imported": count}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/backtest-scorer/experiments/{name}")
async def delete_backtest_experiment(name: str):
    """Delete an experiment."""
    from app.services.backtest_scorer import get_backtest_scorer
    scorer = get_backtest_scorer()
    if scorer.delete_experiment(name):
        return {"status": "deleted"}
    raise HTTPException(status_code=404, detail=f"Experiment '{name}' not found")


@router.get("/backtest-scorer/playbook")
async def get_backtest_playbook():
    """Get the playbook — all proven winning configurations."""
    from app.services.backtest_scorer import get_backtest_scorer
    scorer = get_backtest_scorer()
    return scorer.get_playbook()


@router.post("/backtest-scorer/quick-compare")
async def quick_compare(body: dict):
    """One-shot comparison without creating an experiment.

    Body: {baseline_values: [...], variant_values: [...], baseline_label?, variant_label?}
    """
    from app.services.backtest_scorer import get_backtest_scorer
    scorer = get_backtest_scorer()
    result = scorer.quick_compare(
        baseline_values=body["baseline_values"],
        variant_values=body["variant_values"],
        baseline_label=body.get("baseline_label", "Baseline"),
        variant_label=body.get("variant_label", "Variant"),
    )
    return result
