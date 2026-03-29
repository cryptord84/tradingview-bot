#!/Users/clawbot/Documents/Claude/Projects/tradingview-bot/venv/bin/python3
"""
TradingView Backtest Analyzer
=============================
Export from TradingView Strategy Tester:
  Strategy Tester → any tab → Export icon (bottom right) → saves .xlsx

Then run:
  python tools/analyze_backtest.py path/to/exported.xlsx
  python tools/analyze_backtest.py path/to/exported.csv   (legacy)

Outputs a formatted report you can paste directly to Claude for review.
"""

import sys
import csv
from pathlib import Path
from datetime import datetime


def pf(val):
    """Parse a cell value to float."""
    if val is None:
        return None
    s = str(val).strip()
    if s in ("", "—", "-", "N/A", "None"):
        return None
    try:
        return float(s.replace(",", "").replace("%", "").replace("$", "").strip())
    except ValueError:
        return None


def sheet_to_dict(ws):
    """Convert a two-column key/value sheet to a dict."""
    d = {}
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None and len(row) > 1:
            d[str(row[0]).strip().lower()] = row[1]
    return d


def sheet_to_multidict(ws):
    """Convert a multi-column sheet (row[0]=label, rest=values) to dict."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    d = {}
    for row in rows[1:]:
        if row[0] is None:
            continue
        key = str(row[0]).strip().lower()
        d[key] = {headers[i]: row[i] for i in range(1, len(headers)) if headers[i]}
    return d


def parse_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {name.lower(): wb[name] for name in wb.sheetnames}

    # ── Properties ──────────────────────────────────────────────────────────
    props = {}
    if "properties" in sheets:
        props = sheet_to_dict(sheets["properties"])

    # ── Performance summary ──────────────────────────────────────────────────
    perf = {}
    if "performance" in sheets:
        perf = sheet_to_multidict(sheets["performance"])

    # ── Trades analysis ──────────────────────────────────────────────────────
    ta = {}
    if "trades analysis" in sheets:
        ta = sheet_to_multidict(sheets["trades analysis"])

    # ── Risk-adjusted performance ────────────────────────────────────────────
    risk = {}
    if "risk-adjusted performance" in sheets:
        risk = sheet_to_multidict(sheets["risk-adjusted performance"])

    # ── List of trades ───────────────────────────────────────────────────────
    trades = []
    if "list of trades" in sheets:
        ws = sheets["list of trades"]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [str(c).strip().lower() if c else "" for c in rows[0]]
            # Collect exit rows only (each trade # has an entry+exit pair;
            # the exit row carries the completed trade P&L)
            for row in rows[1:]:
                if row[0] is None:
                    continue
                d = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
                trade_type = str(d.get("type", "")).lower()
                if "exit" in trade_type or "close" in trade_type:
                    trades.append(d)

    return {"props": props, "perf": perf, "ta": ta, "risk": risk, "trades": trades}


def parse_csv_legacy(path):
    """Fallback for plain CSV exports."""
    trades = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(2048)
        f.seek(0)
        delimiter = "," if sample.count(",") > sample.count(";") else ";"
        reader = csv.DictReader(f, delimiter=delimiter)
        for row in reader:
            normalized = {k.strip().lower(): v.strip() for k, v in row.items() if k}
            trades.append(normalized)
    return {"props": {}, "perf": {}, "ta": {}, "risk": {}, "trades": trades}


def get(d, *keys):
    """Safe nested dict get."""
    for k in keys:
        if not isinstance(d, dict):
            return None
        d = d.get(k)
    return d


def build_report(data, file_path):
    props  = data["props"]
    perf   = data["perf"]
    ta     = data["ta"]
    risk   = data["risk"]
    trades = data["trades"]

    if not trades:
        return "ERROR: No completed trades found in the file.\n"

    # ── From summary sheets (authoritative) ──────────────────────────────────
    symbol    = props.get("symbol", "N/A")
    timeframe = props.get("timeframe", "N/A")
    tv_range  = props.get("trading range", props.get("backtesting range", "N/A"))

    net_profit_usd = pf(get(perf, "net profit", "all usd")) or 0
    net_profit_pct = pf(get(perf, "net profit", "all %"))   or 0
    gross_profit   = pf(get(perf, "gross profit", "all usd")) or 0
    gross_loss     = abs(pf(get(perf, "gross loss", "all usd")) or 0)
    max_dd_usd     = pf(get(perf, "max equity drawdown (close-to-close)", "all usd")) or 0
    max_dd_ib      = pf(get(perf, "max equity drawdown (intrabar)", "all usd")) or 0
    expected_payoff= pf(get(perf, "expected payoff", "all usd")) or 0
    bnh_return     = pf(get(perf, "buy & hold return", "all usd")) or 0
    commission     = pf(get(perf, "commission paid", "all usd")) or 0
    init_capital   = pf(get(perf, "initial capital", "all usd")) or 10000

    total_trades   = int(pf(get(ta, "total trades", "all usd")) or len(trades))
    winning_trades = int(pf(get(ta, "winning trades", "all usd")) or 0)
    losing_trades  = int(pf(get(ta, "losing trades", "all usd")) or 0)
    win_rate_all   = pf(get(ta, "percent profitable", "all %")) or 0

    long_total  = int(pf(get(ta, "total trades", "long usd")) or 0)
    long_wins   = int(pf(get(ta, "winning trades", "long usd")) or 0)
    long_wr     = pf(get(ta, "percent profitable", "long %")) or 0
    short_total = int(pf(get(ta, "total trades", "short usd")) or 0)
    short_wins  = int(pf(get(ta, "winning trades", "short usd")) or 0)
    short_wr    = pf(get(ta, "percent profitable", "short %")) or 0

    profit_factor = pf(get(risk, "profit factor", "all usd")) or 0
    sharpe        = pf(get(risk, "sharpe ratio", "all usd"))
    sortino       = pf(get(risk, "sortino ratio", "all usd"))

    long_pf  = pf(get(risk, "profit factor", "long usd"))
    short_pf = pf(get(risk, "profit factor", "short usd"))

    # ── From individual trades ────────────────────────────────────────────────
    profits = []
    for t in trades:
        p = pf(t.get("net p&l usd")) or pf(t.get("profit")) or pf(t.get("net p&l"))
        if p is not None:
            profits.append(p)

    wins_list   = [p for p in profits if p > 0]
    losses_list = [p for p in profits if p <= 0]
    avg_win     = sum(wins_list)   / len(wins_list)   if wins_list   else 0
    avg_loss    = abs(sum(losses_list) / len(losses_list)) if losses_list else 0
    largest_win  = max(wins_list)   if wins_list   else 0
    largest_loss = min(losses_list) if losses_list else 0

    # Consecutive streaks
    max_cw = max_cl = cw = cl = 0
    for p in profits:
        if p > 0:
            cw += 1; cl = 0
        else:
            cl += 1; cw = 0
        max_cw = max(max_cw, cw)
        max_cl = max(max_cl, cl)

    # Cumulative equity curve
    cum_curve = []
    c = 0.0
    for p in profits:
        c += p
        cum_curve.append(c)

    # Date range from trades
    dates = []
    for t in trades:
        dt = t.get("date and time") or t.get("date") or t.get("date/time")
        if dt:
            dates.append(str(dt))
    date_range = tv_range if tv_range != "N/A" else (
        f"{dates[0]}  →  {dates[-1]}" if len(dates) >= 2 else "N/A"
    )

    # Long/short P&L from trades
    long_pnl  = sum(pf(t.get("net p&l usd")) or 0 for t in trades
                    if "long" in str(t.get("type", "")).lower() or "buy" in str(t.get("signal", "")).lower())
    short_pnl = sum(pf(t.get("net p&l usd")) or 0 for t in trades
                    if "short" in str(t.get("type", "")).lower() or "sell" in str(t.get("signal", "")).lower())

    # Best/worst 5 individual trades
    trade_details = []
    for t in trades:
        p = pf(t.get("net p&l usd")) or pf(t.get("profit")) or 0
        dt = str(t.get("date and time") or t.get("date") or "")
        sig = str(t.get("signal") or t.get("type") or "")
        fav = pf(t.get("favorable excursion usd")) or 0
        adv = pf(t.get("adverse excursion usd")) or 0
        trade_details.append({"profit": p, "date": dt, "signal": sig, "runup": fav, "drawdown": adv})

    best5  = sorted(trade_details, key=lambda x: x["profit"], reverse=True)[:5]
    worst5 = sorted(trade_details, key=lambda x: x["profit"])[:5]

    # ── Build report ─────────────────────────────────────────────────────────
    lines = []
    a = lines.append

    a("=" * 66)
    a("  TRADINGVIEW BACKTEST REPORT")
    a(f"  File      : {Path(file_path).name}")
    a(f"  Symbol    : {symbol}  |  Timeframe: {timeframe}")
    a(f"  Period    : {date_range}")
    a(f"  Capital   : ${init_capital:,.2f}")
    a("=" * 66)

    a("")
    a("── PERFORMANCE SUMMARY ──────────────────────────────────────────")
    a(f"  Net Profit          : ${net_profit_usd:+,.2f}  ({net_profit_pct:+.2f}%)")
    a(f"  Gross Profit        : ${gross_profit:,.2f}")
    a(f"  Gross Loss          : -${gross_loss:,.2f}")
    a(f"  Profit Factor       : {profit_factor:.3f}")
    a(f"  Buy & Hold Return   : ${bnh_return:+,.2f}  (strategy {'outperformed' if net_profit_usd > bnh_return else 'underperformed'})")
    a(f"  Commission Paid     : ${commission:,.2f}")

    a("")
    a("── TRADES ───────────────────────────────────────────────────────")
    a(f"  Total Trades        : {total_trades}")
    a(f"  Winners / Losers    : {winning_trades} / {losing_trades}")
    a(f"  Win Rate            : {win_rate_all:.2f}%")
    a(f"  Expected Payoff     : ${expected_payoff:+.2f} per trade")

    a("")
    a("── RISK / REWARD ────────────────────────────────────────────────")
    a(f"  Avg Win             : ${avg_win:,.2f}")
    a(f"  Avg Loss            : -${avg_loss:,.2f}")
    a(f"  Win/Loss Ratio      : {(avg_win/avg_loss):.2f}" if avg_loss else "  Win/Loss Ratio      : N/A")
    a(f"  Largest Win         : ${largest_win:+,.2f}")
    a(f"  Largest Loss        : ${largest_loss:+,.2f}")
    a(f"  Max Drawdown (EOB)  : -${max_dd_usd:,.2f}")
    a(f"  Max Drawdown (IB)   : -${max_dd_ib:,.2f}")

    a("")
    a("── RISK-ADJUSTED ────────────────────────────────────────────────")
    a(f"  Sharpe Ratio        : {sharpe:.3f}" if sharpe is not None else "  Sharpe Ratio        : N/A")
    a(f"  Sortino Ratio       : {sortino:.3f}" if sortino is not None else "  Sortino Ratio       : N/A")

    a("")
    a("── LONG vs SHORT ────────────────────────────────────────────────")
    a(f"  Long  : {long_total:>3} trades  |  {long_wins} wins ({long_wr:.1f}%)  |  P&L: ${long_pnl:+,.2f}  |  PF: {long_pf:.3f}" if long_pf else
      f"  Long  : {long_total:>3} trades  |  {long_wins} wins ({long_wr:.1f}%)  |  P&L: ${long_pnl:+,.2f}")
    a(f"  Short : {short_total:>3} trades  |  {short_wins} wins ({short_wr:.1f}%)  |  P&L: ${short_pnl:+,.2f}  |  PF: {short_pf:.3f}" if short_pf else
      f"  Short : {short_total:>3} trades  |  {short_wins} wins ({short_wr:.1f}%)  |  P&L: ${short_pnl:+,.2f}")

    a("")
    a("── STREAKS ──────────────────────────────────────────────────────")
    a(f"  Max Consec Wins     : {max_cw}")
    a(f"  Max Consec Losses   : {max_cl}")

    a("")
    a("── BEST 5 TRADES ────────────────────────────────────────────────")
    for i, t in enumerate(best5, 1):
        a(f"  {i}. ${t['profit']:+8.2f}  {t['date'][:19]:<20}  {t['signal']}")

    a("")
    a("── WORST 5 TRADES ───────────────────────────────────────────────")
    for i, t in enumerate(worst5, 1):
        a(f"  {i}. ${t['profit']:+8.2f}  {t['date'][:19]:<20}  {t['signal']}")

    a("")
    a("── ALL TRADES (Trade # | P&L | Signal | Date) ───────────────────")
    for i, t in enumerate(trade_details, 1):
        a(f"  {i:>2}. ${t['profit']:+8.2f}  {t['date'][:19]:<20}  {t['signal']}")

    a("")
    a("── EQUITY CURVE (ASCII) ─────────────────────────────────────────")
    if cum_curve:
        width  = 50
        mn, mx = min(cum_curve), max(cum_curve)
        rng    = mx - mn if mx != mn else 1
        step   = max(1, len(cum_curve) // width)
        sampled = cum_curve[::step][:width]
        bars   = "▁▂▃▄▅▆▇█"
        spark  = "".join(bars[min(7, int((v - mn) / rng * 8))] for v in sampled)
        a(f"  {spark}")
        a(f"  start: $0.00   min: ${mn:+,.2f}   max: ${mx:+,.2f}   final: ${cum_curve[-1]:+,.2f}")

    a("")
    a("=" * 66)
    a("  Paste this report to Claude for strategy analysis.")
    a("=" * 66)

    return "\n".join(lines)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nUsage: python tools/analyze_backtest.py <path/to/export.xlsx>")
        sys.exit(1)

    file_path = sys.argv[1]
    if not Path(file_path).exists():
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)

    if Path(file_path).suffix.lower() in (".xlsx", ".xls"):
        data = parse_xlsx(file_path)
    else:
        raw, _ = parse_csv_legacy(file_path)
        data = {"props": {}, "perf": {}, "ta": {}, "risk": {}, "trades": raw}

    report = build_report(data, file_path)
    print(report)

    out_path = Path(file_path).with_name(Path(file_path).stem + ".report.txt")
    out_path.write_text(report)
    print(f"\nReport saved to: {out_path}")


if __name__ == "__main__":
    main()
